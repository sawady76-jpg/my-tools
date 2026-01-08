import pandas as pd
import glob
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ==========================================
# ★追加：実行場所をスクリプトのあるフォルダに強制変更
# ==========================================
try:
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
except:
    pass

# ==========================================
# 設定エリア
# ==========================================
INPUT_FILE_PATTERN = "集計結果_*.xlsx"
TEMPLATE_FILENAME = "集計用テンプレート_v1.xlsx"

# ==========================================
# 1. データの読み込みとクリーニング関数
# ==========================================
def load_and_clean_data_from_excel(file_path):
    print(f"読み込み中: {file_path}")
    data_dict = {}

    try:
        xls = pd.ExcelFile(file_path)
    except Exception as e:
        print(f"エラー: ファイルを開けませんでした。 {e}")
        return {}

    sheet_names = {
        1: "1.着信件数",
        2: "2.従業員別",
        3: "3.関数_拠点別",
        4: "4.時短勤務",
        5: "5.営業時間内集計",
        6: "6.時間内集計"
    }

    # --- 1. 着信件数 ---
    if sheet_names[1] in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_names[1], header=None)
        start_row = 0
        for i, row in df.iterrows():
            if str(row[0]) == "拠点":
                start_row = i + 1
                break
        
        if start_row > 0:
            df_clean = df.iloc[start_row:, 0:7].copy()
            df_clean.columns = ["拠点", "内線_入電", "内線_着電", "外線_入電", "外線_着電", "他拠点へ転送", "他拠点から転送"]
            df_clean = df_clean.dropna(subset=["拠点"])
            for c in ["内線_入電", "内線_着電", "外線_入電", "外線_着電"]:
                df_clean[c] = pd.to_numeric(df_clean[c], errors='coerce').fillna(0).astype(int)
            data_dict[1] = df_clean
            print("  OK: シート1 (着信件数)")

    # --- 2. 従業員別 ---
    if sheet_names[2] in xls.sheet_names:
        data_dict[2] = pd.read_excel(xls, sheet_name=sheet_names[2])
        print("  OK: シート2 (従業員別)")

    # --- 3. 関数_拠点別 ---
    if sheet_names[3] in xls.sheet_names:
        data_dict[3] = pd.read_excel(xls, sheet_name=sheet_names[3])
        print("  OK: シート3 (関数_拠点別)")

    # --- 4. 時短勤務 ---
    if sheet_names[4] in xls.sheet_names:
        data_dict[4] = pd.read_excel(xls, sheet_name=sheet_names[4])
        print("  OK: シート4 (時短勤務)")

    # --- 5. 営業時間内集計 ---
    if sheet_names[5] in xls.sheet_names:
        # 注釈行があるため、headerの位置に注意（直前のツールで注釈を入れたため header=1 か 2 になる）
        # 安全のため "拠点名" を探す
        df_raw = pd.read_excel(xls, sheet_name=sheet_names[5], header=None)
        header_idx = 0
        for i, row in df_raw.iterrows():
            if "拠点名" in str(row.values):
                header_idx = i
                break
        
        df = pd.read_excel(xls, sheet_name=sheet_names[5], header=header_idx)
        target_cols = ["拠点名", "営業時間内_外線のみ", "人員", "1人当たり／月", "全体からの比率"]
        valid_cols = [c for c in target_cols if c in df.columns]
        data_dict[5] = df[valid_cols]
        print("  OK: シート5 (営業時間内集計)")

    # --- 6. 時間内集計 ---
    if sheet_names[6] in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sheet_names[6], header=None)
        start_row = 0
        for i, row in df.iterrows():
            if str(row[0]) == "拠点":
                start_row = i + 1
                break
        
        if start_row > 0:
            df_clean = df.iloc[start_row:, 0:7].copy()
            df_clean.columns = ["拠点", "内線_入電", "内線_着電", "内線_応答率", "外線_入電", "外線_着電", "外線_応答率"]
            df_clean = df_clean.dropna(subset=["拠点"])
            data_dict[6] = df_clean
            print("  OK: シート6 (時間内集計)")

    return data_dict

# ==========================================
# 2. Excelテンプレート作成関数
# ==========================================
def create_template_excel(data_dict):
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    configs = [
        (1, "着信", "1. 拠点別 着信件数"),
        (2, "従業員", "2. 従業員別 実績"),
        (3, "関数拠点", "3. 拠点別 関数集計"),
        (4, "時短", "4. 時短勤務者"),
        (5, "営業集計", "5. 営業時間内(Final) 集計"),
        (6, "時間内", "6. 営業時間内(Target) 集計")
    ]

    for key, sheet_base_name, title in configs:
        if key not in data_dict:
            continue
        
        df = data_dict[key]
        
        # Excelシート作成
        ws_report = wb.create_sheet(title=f"Report_{sheet_base_name}")
        ws_raw = wb.create_sheet(title=f"RawData_{sheet_base_name}")

        # RawData準備
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            ws_raw.cell(row=1, column=col_idx, value=header)
        
        # データ書き込み (RawData)
        for r_idx, row in enumerate(df.itertuples(index=False), 2):
            for c_idx, value in enumerate(row, 1):
                ws_raw.cell(row=r_idx, column=c_idx, value=value)

        # Report準備
        ws_report["A1"] = title
        ws_report["A1"].font = Font(size=14, bold=True)

        for col_idx, header in enumerate(headers, 1):
            cell = ws_report.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=Side(style='thin'))
            ws_report.column_dimensions[get_column_letter(col_idx)].width = 15

        start_row = 4
        num_rows = 50 
        for r in range(start_row, start_row + num_rows):
            for c in range(1, len(headers) + 1):
                col_letter = get_column_letter(c)
                raw_row = r - start_row + 2
                raw_cell_ref = f"'RawData_{sheet_base_name}'!{col_letter}{raw_row}"
                formula = f'=IF({raw_cell_ref}="","", {raw_cell_ref})'
                cell = ws_report.cell(row=r, column=c)
                cell.value = formula
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), bottom=Side(style='dotted'))

    wb.save(TEMPLATE_FILENAME)
    print(f"\n完了！\n -> {TEMPLATE_FILENAME} を作成しました。")

# ==========================================
# メイン処理
# ==========================================
if __name__ == "__main__":
    # 最新の集計結果Excelを探す
    files = glob.glob(INPUT_FILE_PATTERN)
    if not files:
        print("エラー: '集計結果_*.xlsx' が見つかりません。")
        print("注意: まず 'analyze_logs.py' を実行して、集計結果ファイルを作成してください。")
        input("Enterキーを押して終了...")
    else:
        # 作成日時が新しい順に並べて最新を取得
        latest_file = max(files, key=os.path.getctime)
        print(f"最新のファイルを使用します: {latest_file}")
        
        cleaned_data = load_and_clean_data_from_excel(latest_file)
        if cleaned_data:
            create_template_excel(cleaned_data)
        
        input("Enterキーを押して終了...")