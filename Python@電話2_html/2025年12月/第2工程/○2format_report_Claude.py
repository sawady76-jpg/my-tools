#!/usr/bin/env python3
"""
集計結果Excelファイルの書式を整えるPythonスクリプト
入力: generate_report_Claude.pyで生成されたExcelファイル
出力: 書式が整えられたExcelファイル

使用方法:
    python format_report.py 集計結果_2026-01-08.xlsx
    または
    python format_report.py  （カレントディレクトリの最新の集計結果ファイルを処理）
"""

import sys
import os
import glob
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# 書式設定
# ============================================================

# フォント設定
FONT_NAME = 'BIZ UDゴシック'
FONT_SIZE = 11
FONT_SIZE_NOTE = 10

# 色設定
HEADER_BG_COLOR = 'FF1F497D'  # ヘッダー背景色（濃紺）
HEADER_FONT_COLOR = 'FFFFFFFF'  # ヘッダー文字色（白）
DATA_BG_COLOR = 'FFFFFFFF'  # データ背景色（白）

# 罫線設定
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def apply_header_style(cell):
    """ヘッダーセルのスタイルを適用"""
    cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=HEADER_FONT_COLOR)
    cell.fill = PatternFill(start_color=HEADER_BG_COLOR, end_color=HEADER_BG_COLOR, fill_type='solid')
    cell.border = THIN_BORDER
    cell.alignment = Alignment(horizontal='center', vertical='center')


def apply_data_style(cell):
    """データセルのスタイルを適用"""
    cell.font = Font(name=FONT_NAME, size=FONT_SIZE)
    cell.fill = PatternFill(start_color=DATA_BG_COLOR, end_color=DATA_BG_COLOR, fill_type='solid')
    cell.border = THIN_BORDER


def apply_note_style(cell):
    """注釈セルのスタイルを適用"""
    cell.font = Font(name=FONT_NAME, size=FONT_SIZE_NOTE, bold=True)


def format_sheet1(ws):
    """シート1: 着信件数の書式設定"""
    print("  シート1: 着信件数を書式設定中...")
    
    # 列幅設定
    col_widths = {'A': 13.2, 'B': 6.5, 'C': 6.5, 'D': 7.7, 'E': 7.7, 'F': 8.8, 'G': 9.9, 'H': 2.0,
                  'J': 18.0, 'K': 14.0, 'L': 10.0, 'M': 10.0}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # 既存の結合を解除（再設定のため）
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))
    
    # セル結合
    ws.merge_cells('A1:A2')  # 拠点
    ws.merge_cells('B1:C1')  # 内線
    ws.merge_cells('D1:E1')  # 外線
    ws.merge_cells('F1:F2')  # 他拠点へ転送
    ws.merge_cells('G1:G2')  # 他拠点から転送
    
    # ヘッダーの値を設定
    ws['A1'] = '拠点'
    ws['B1'] = '内線'
    ws['D1'] = '外線'
    ws['F1'] = '他拠点へ転送'
    ws['G1'] = '他拠点から転送'
    ws['B2'] = '入電'
    ws['C2'] = '着電'
    ws['D2'] = '入電'
    ws['E2'] = '着電'
    
    # ヘッダー行（1行目、2行目）のスタイル
    header_cells = ['A1', 'B1', 'C1', 'D1', 'E1', 'F1', 'G1', 'A2', 'B2', 'C2', 'D2', 'E2', 'F2', 'G2']
    for cell_ref in header_cells:
        cell = ws[cell_ref]
        cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=HEADER_FONT_COLOR)
        cell.fill = PatternFill(start_color=HEADER_BG_COLOR, end_color=HEADER_BG_COLOR, fill_type='solid')
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # データ行（3行目以降）
    for row in range(3, ws.max_row + 1):
        for col in range(1, 8):  # A〜G列
            cell = ws.cell(row=row, column=col)
            apply_data_style(cell)
            if col == 1:  # 拠点名は左寄せ
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:  # 数値は右寄せ
                cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # 追加集計部分（J列以降）
    if ws['J26'].value:
        apply_note_style(ws['J26'])
    for row in range(27, 37):
        for col in range(10, 14):  # J〜M列
            cell = ws.cell(row=row, column=col)
            if row == 27:
                apply_header_style(cell)
            elif cell.value is not None:
                apply_data_style(cell)


def format_sheet2(ws):
    """シート2: 従業員別の書式設定"""
    print("  シート2: 従業員別を書式設定中...")
    
    # 列幅設定
    ws.column_dimensions['A'].width = 16.1
    ws.column_dimensions['B'].width = 10.0
    ws.column_dimensions['C'].width = 6.5
    ws.column_dimensions['D'].width = 5.4
    ws.column_dimensions['E'].width = 8.8
    ws.column_dimensions['F'].width = 6.5
    ws.column_dimensions['G'].width = 9.9
    
    # 日付列（H列以降）
    for col in range(8, ws.max_column - 2):
        ws.column_dimensions[get_column_letter(col)].width = 13.2
    
    # 最後の3列
    ws.column_dimensions[get_column_letter(ws.max_column - 2)].width = 5.4
    ws.column_dimensions[get_column_letter(ws.max_column - 1)].width = 10.0
    ws.column_dimensions[get_column_letter(ws.max_column)].width = 9.3
    
    # 行高さ（ヘッダー行）
    ws.row_dimensions[1].height = 25.2
    
    # ヘッダー行
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        apply_header_style(cell)
    
    # データ行
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            apply_data_style(cell)
            if col <= 3:  # 名前、拠点、受発注は左寄せ
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:  # 数値は右寄せ
                cell.alignment = Alignment(horizontal='right', vertical='center')


def format_sheet3(ws):
    """シート3: 関数_拠点別の書式設定"""
    print("  シート3: 関数_拠点別を書式設定中...")
    
    # 列幅設定
    col_widths = {'A': 20.4, 'B': 16.5, 'C': 10.0, 'D': 6.0, 'E': 11.0, 'F': 12.0, 
                  'G': 2.0, 'I': 18.0, 'J': 10.0, 'K': 10.0, 'L': 11.0}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # 行高さ
    ws.row_dimensions[1].height = 25.2
    
    # 左側テーブル（A〜F列）
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        apply_header_style(cell)
    
    for row in range(2, ws.max_row + 1):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                apply_data_style(cell)
                if col == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                
                # F列（係数）を%表記に変換
                if col == 6 and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0%'
    
    # 右側テーブル（I〜L列）
    for col in range(9, 13):
        cell = ws.cell(row=1, column=col)
        if cell.value is not None:
            apply_header_style(cell)
    
    for row in range(2, ws.max_row + 1):
        for col in range(9, 13):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                apply_data_style(cell)
                if col == 9:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center')


def format_sheet4(ws):
    """シート4: 時短勤務の書式設定"""
    print("  シート4: 時短勤務を書式設定中...")
    
    # 列幅設定
    col_widths = {'A': 13.9, 'B': 20.4, 'C': 9.5, 'D': 8.3, 'E': 11.7, 'F': 11.0}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # 行高さ
    ws.row_dimensions[1].height = 25.2
    
    # ヘッダー行
    for col in range(1, 7):
        cell = ws.cell(row=1, column=col)
        apply_header_style(cell)
    
    # データ行
    for row in range(2, ws.max_row + 1):
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            apply_data_style(cell)
            if col <= 2:  # 氏名、部署は左寄せ
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:  # 数値は右寄せ
                cell.alignment = Alignment(horizontal='right', vertical='center')


def format_sheet5(ws):
    """シート5: 営業時間内集計の書式設定"""
    print("  シート5: 営業時間内集計を書式設定中...")
    
    # 列幅設定
    col_widths = {'A': 13.2, 'B': 14.0, 'C': 4.4, 'D': 9.9, 'E': 25.3}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # 注釈行
    apply_note_style(ws['A1'])
    
    # 行高さ
    ws.row_dimensions[2].height = 25.2
    
    # ヘッダー行（2行目）
    for col in range(1, 6):
        cell = ws.cell(row=2, column=col)
        apply_header_style(cell)
    
    # データ行
    for row in range(3, ws.max_row + 1):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            apply_data_style(cell)
            if col == 1:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='right', vertical='center')
            
            # E列（係数など）を%表記に変換
            if col == 5 and isinstance(cell.value, (int, float)):
                cell.number_format = '0.0%'


def format_sheet6(ws):
    """シート6: 時間内集計の書式設定"""
    print("  シート6: 時間内集計を書式設定中...")
    
    # 列幅設定
    col_widths = {'A': 13.2, 'B': 6.5, 'C': 6.5, 'D': 10.0, 'E': 7.7, 'F': 6.5, 'G': 10.0,
                  'H': 6.5, 'M': 18.3, 'N': 9.5, 'O': 9.4, 'P': 10.0}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # 注釈行
    apply_note_style(ws['A1'])
    
    # 既存の結合を解除（再設定のため）
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))
    
    # セル結合（ヘッダー部分）
    ws.merge_cells('A3:A5')  # 拠点
    ws.merge_cells('B3:D4')  # 内線
    ws.merge_cells('E3:G4')  # 外線
    
    # ヘッダーの値を設定
    ws['A3'] = '拠点'
    ws['B3'] = '内線'
    ws['E3'] = '外線'
    ws['B5'] = '入電'
    ws['C5'] = '着電'
    ws['D5'] = '応答率'
    ws['E5'] = '入電'
    ws['F5'] = '着電'
    ws['G5'] = '応答率'
    
    # ヘッダーセルのスタイル（3〜5行目、A〜G列）
    for row in range(3, 6):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color=HEADER_FONT_COLOR)
            cell.fill = PatternFill(start_color=HEADER_BG_COLOR, end_color=HEADER_BG_COLOR, fill_type='solid')
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # メインデータ（6行目以降、A〜G列）
    for row in range(6, ws.max_row + 1):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                apply_data_style(cell)
                if col == 1:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                
                # 応答率列（D列とG列）を%表記に変換
                if col in [4, 7] and isinstance(cell.value, (int, float)) and cell.value != 0:
                    cell.number_format = '0.0%'
    
    # 右側の追加集計
    # 全体集計ヘッダー（5行目）
    right_headers = ['M5', 'N5', 'O5', 'P5', 'Q5']
    for cell_ref in right_headers:
        cell = ws[cell_ref]
        if cell.value:
            apply_header_style(cell)
    
    # 全体集計データ（6行目）
    for col in range(13, 18):
        cell = ws.cell(row=6, column=col)
        if cell.value is not None:
            apply_data_style(cell)
            # Q列（全応答率）を%表記に
            if col == 17 and isinstance(cell.value, (int, float)):
                cell.number_format = '0.0%'
    
    # 追加集計ヘッダー（9行目）
    if ws['M9'].value:
        apply_note_style(ws['M9'])
    for col in range(14, 17):
        cell = ws.cell(row=9, column=col)
        if cell.value:
            apply_header_style(cell)
    
    # 追加集計データ（10行目以降）
    for row in range(10, 20):
        for col in range(13, 17):
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                apply_data_style(cell)
                # P列（％）を%表記に
                if col == 16 and isinstance(cell.value, (int, float)):
                    cell.number_format = '0.0%'


def format_excel(filepath):
    """Excelファイル全体の書式を設定"""
    print(f"\n書式設定中: {filepath}")
    
    wb = load_workbook(filepath)
    
    # 各シートの書式設定
    if '1.着信件数' in wb.sheetnames:
        format_sheet1(wb['1.着信件数'])
    
    if '2.従業員別' in wb.sheetnames:
        format_sheet2(wb['2.従業員別'])
    
    if '3.関数_拠点別' in wb.sheetnames:
        format_sheet3(wb['3.関数_拠点別'])
    
    if '4.時短勤務' in wb.sheetnames:
        format_sheet4(wb['4.時短勤務'])
    
    if '5.営業時間内集計' in wb.sheetnames:
        format_sheet5(wb['5.営業時間内集計'])
    
    if '6.時間内集計' in wb.sheetnames:
        format_sheet6(wb['6.時間内集計'])
    
    # 保存
    wb.save(filepath)
    print(f"\n完了！書式設定済みファイル: {filepath}")


def main():
    # コマンドライン引数からファイルパスを取得
    if len(sys.argv) > 1:
        filepath = sys.argv[1]
    else:
        # 引数がない場合は最新の集計結果ファイルを探す
        files = glob.glob('集計結果_*.xlsx')
        if not files:
            files = glob.glob('/mnt/user-data/outputs/集計結果_*.xlsx')
        
        if not files:
            print("エラー: 集計結果ファイルが見つかりません。")
            print("使用方法: python format_report.py <ファイルパス>")
            sys.exit(1)
        
        filepath = max(files, key=os.path.getctime)
        print(f"最新のファイルを処理します: {filepath}")
    
    if not os.path.exists(filepath):
        print(f"エラー: ファイルが見つかりません: {filepath}")
        sys.exit(1)
    
    format_excel(filepath)


if __name__ == '__main__':
    main()
