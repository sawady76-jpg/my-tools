#!/usr/bin/env python3
"""
発着信履歴統合版から集計結果を生成するPythonスクリプト
入力: 発着信履歴_統合版.xlsx
出力: 集計結果_YYYY-MM-DD.xlsx（理想_集計結果と同形式）
"""

import pandas as pd
import numpy as np
import re
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================
# 設定
# ============================================================
INPUT_FILE = '/mnt/user-data/uploads/発着信履歴_統合版.xlsx'
OUTPUT_FILE = '/mnt/user-data/outputs/集計結果_{}.xlsx'.format(datetime.now().strftime('%Y-%m-%d'))

# 拠点の順序（出力見本に合わせる）
BASE_ORDER = [
    '東京', '横浜', '埼玉', '滋賀', '大阪', '千葉', '福岡', '岡山', '名古屋', '仙台',
    '流山', '広島', '本社輸入', '本社在庫物流課', '本社購買', '静岡', 'MF東京', '本社',
    '本社経理', '本社EC', '本社総務', '本社商品開発', 'MF大阪', '本社営業本部', '多摩',
    '大阪(ダイヤテック)', '本社IT', '北関東', '札幌', '熊本', '本社取締役', '金沢'
]

# 受発注担当者リスト（出力見本から抽出）
JUHATCHU_LIST = [
    '佐々木美咲', '北林友希', '坂田智世', '林まど佳', '西垣彩', '高澤早紀',
    '中山満里奈', '奥秋素子', '山﨑啓右', '川崎瞳', '舞智江', '萩原直美', '青木観奈',
    '池田千恵子', '矢島静音', '石川恵理', '竹内梓', '細田昌子', '豊田里花', '金子友希',
    '北出智子', '松井治美', '梅原薫', '筆坂雪子',
    '松下愛',
    '小島佳菜',
    '奥薗美和', '山崎芹奈', '松尾明日香', '松添愛', '重松宝成',
    '伊藤優', '岩佐寛子', '岸本麻衣子', '武政彩果', '河原由布子', '矢野夏絵',
    '大渕温子', '水島奈美', '玉腰千恵', '稲垣みちる',
    '仁藤佳美', '真壁彰子', '阿部かおり', '齋藤詩織',
    '鴻池千紗都',
]

# 時短勤務者データ（見本から直接設定）
JITAN_DATA = [
    {'氏名': '玉腰　千恵', '部署': '名古屋営業所(業務)', '勤務時間': 5.8, '係数': 1.39, '名前_key': '玉腰千恵'},
    {'氏名': '石川　恵理', '部署': '埼玉支店(業務)', '勤務時間': 6.3, '係数': 1.28, '名前_key': '石川恵理'},
    {'氏名': '矢島　静音', '部署': '埼玉支店(業務)', '勤務時間': 6.0, '係数': 1.33, '名前_key': '矢島静音'},
    {'氏名': '河原　由布子', '部署': '岡山営業所(業務)', '勤務時間': 6.0, '係数': 1.33, '名前_key': '河原由布子'},
]

# 転送関係
TRANSFER_TO = {
    '大阪': '岡山', '千葉': '埼玉', '流山': '東京', '広島': '福岡', '静岡': '岡山',
    '多摩': '横浜', '北関東': '埼玉', '札幌': '仙台', '熊本': '福岡', '金沢': '滋賀'
}
TRANSFER_FROM = {
    '東京': '流山', '横浜': '多摩', '埼玉': '千葉・北関東', '滋賀': '金沢',
    '福岡': '広島・熊本', '岡山': '大阪・静岡', '仙台': '札幌'
}

# 地域グループ
REGION_GROUPS = {
    '東京+流山': ['東京', '流山'],
    '埼玉+千葉+北関東': ['埼玉', '千葉', '北関東'],
    '横浜+多摩': ['横浜', '多摩'],
    '仙台+札幌': ['仙台', '札幌'],
    '滋賀+金沢': ['滋賀', '金沢'],
    '名古屋': ['名古屋'],
    '福岡+熊本+広島': ['福岡', '熊本', '広島'],
    '岡山+大阪+静岡': ['岡山', '大阪', '静岡'],
}

# 受発注集計用の拠点
JUHATCHU_BASE_ORDER = ['仙台', '千葉', '名古屋', '埼玉', '大阪', '岡山', '東京', '横浜', '流山', '滋賀', '福岡']

# 除外する名前
EXCLUDE_NAMES = {'受付用電話機', '大阪出張用', '受付電話', '廣田優希', '相川大毅', '吉田夏子', '畑中博孝'}


# ============================================================
# ユーティリティ関数
# ============================================================
def extract_info(name_str):
    """【拠点】名前 形式から拠点と名前を抽出"""
    if pd.isna(name_str) or name_str == '不在':
        return None, None
    match = re.match(r'【(.+?)】(.+)', str(name_str))
    if match:
        return match.group(1), match.group(2)
    return None, str(name_str)


def is_business_hours(row):
    """営業時間内かどうかを判定（8:45-17:45）"""
    h, m = row['時間'], row['分']
    if h < 8 or h > 17:
        return False
    if h == 8 and m < 45:
        return False
    if h == 17 and m >= 45:
        return False
    return True


# ============================================================
# データ読み込みと前処理
# ============================================================
print("データを読み込み中...")

df_naisen = pd.read_excel(INPUT_FILE, sheet_name='内線通話')
df_gaisen_hasshin = pd.read_excel(INPUT_FILE, sheet_name='外線発信')
df_gaisen_chakushin = pd.read_excel(INPUT_FILE, sheet_name='外線着信')

for df in [df_naisen, df_gaisen_hasshin, df_gaisen_chakushin]:
    df['時刻'] = pd.to_datetime(df['時刻'])
    df['日付'] = df['時刻'].dt.date
    df['時間'] = df['時刻'].dt.hour
    df['分'] = df['時刻'].dt.minute

# 拠点と名前を抽出
df_naisen['発信者_拠点'], df_naisen['発信者_名前'] = zip(*df_naisen['発信者'].apply(extract_info))
df_naisen['着信者_拠点'], df_naisen['着信者_名前'] = zip(*df_naisen['着信者'].apply(extract_info))
df_naisen['最終着信者_拠点'], df_naisen['最終着信者_名前'] = zip(*df_naisen['最終着信者'].apply(extract_info))

df_gaisen_hasshin['発信者_拠点'], df_gaisen_hasshin['発信者_名前'] = zip(*df_gaisen_hasshin['発信者'].apply(extract_info))

df_gaisen_chakushin['着信者_拠点'] = df_gaisen_chakushin['着信者'].apply(lambda x: extract_info(x)[0])
df_gaisen_chakushin['最終着信者_拠点'], df_gaisen_chakushin['最終着信者_名前'] = zip(*df_gaisen_chakushin['最終着信者'].apply(extract_info))

# 営業時間フラグ
df_naisen['営業時間内'] = df_naisen.apply(is_business_hours, axis=1)
df_gaisen_chakushin['営業時間内'] = df_gaisen_chakushin.apply(is_business_hours, axis=1)

# 日付範囲（NaTを除外）
all_dates = sorted(set(df_naisen['日付'].dropna()) | set(df_gaisen_hasshin['日付'].dropna()) | set(df_gaisen_chakushin['日付'].dropna()))
date_range = pd.date_range(start=min(all_dates), end=max(all_dates))

print(f"データ期間: {min(all_dates)} ～ {max(all_dates)}")


# ============================================================
# シート1: 着信件数
# ============================================================
print("シート1: 着信件数を集計中...")

naisen_nyuden = df_naisen.groupby('着信者_拠点').size()
naisen_chakuden = df_naisen.groupby('最終着信者_拠点').size()
gaisen_nyuden = df_gaisen_chakushin.groupby('着信者_拠点').size()
gaisen_chakuden = df_gaisen_chakushin.groupby('最終着信者_拠点').size()

sheet1_data = []
for base in BASE_ORDER:
    sheet1_data.append({
        '拠点': base,
        '内線_入電': naisen_nyuden.get(base, 0),
        '内線_着電': naisen_chakuden.get(base, 0),
        '外線_入電': gaisen_nyuden.get(base, 0),
        '外線_着電': gaisen_chakuden.get(base, 0),
        '他拠点へ転送': TRANSFER_TO.get(base, ''),
        '他拠点から転送': TRANSFER_FROM.get(base, ''),
    })

sheet1_data.append({
    '拠点': '合計',
    '内線_入電': naisen_nyuden.sum(),
    '内線_着電': naisen_chakuden.sum(),
    '外線_入電': len(df_gaisen_chakushin),
    '外線_着電': gaisen_chakuden.sum(),
    '他拠点へ転送': '',
    '他拠点から転送': '',
})

# 追加集計（24H）
additional_24h = []
for group_name, bases in REGION_GROUPS.items():
    nyuden = sum(gaisen_nyuden.get(b, 0) for b in bases)
    chakuden = sum(gaisen_chakuden.get(b, 0) for b in bases)
    ratio = chakuden / nyuden if nyuden > 0 else 0
    additional_24h.append({'グループ': group_name, '電話が入った数': nyuden, 'とった数': chakuden, '％': ratio})

total_nyuden = sum(r['電話が入った数'] for r in additional_24h)
total_chakuden = sum(r['とった数'] for r in additional_24h)
additional_24h.append({'グループ': '合計', '電話が入った数': total_nyuden, 'とった数': total_chakuden, 
                       '％': total_chakuden / total_nyuden if total_nyuden > 0 else 0})


# ============================================================
# シート2: 従業員別
# ============================================================
print("シート2: 従業員別を集計中...")

# 各従業員の主要拠点を特定（最終着信者として最も多い拠点）
employee_main_base = {}

# 外線着信ベースで主要拠点を判定（最も重要な指標）
gaisen_by_emp = df_gaisen_chakushin[df_gaisen_chakushin['最終着信者_名前'].notna()].groupby(
    ['最終着信者_名前', '最終着信者_拠点']).size().reset_index(name='count')
for name in gaisen_by_emp['最終着信者_名前'].unique():
    emp_data = gaisen_by_emp[gaisen_by_emp['最終着信者_名前'] == name]
    main_base = emp_data.loc[emp_data['count'].idxmax(), '最終着信者_拠点']
    employee_main_base[name] = main_base

# 内線でも補完
naisen_by_emp = df_naisen[df_naisen['最終着信者_名前'].notna()].groupby(
    ['最終着信者_名前', '最終着信者_拠点']).size().reset_index(name='count')
for name in naisen_by_emp['最終着信者_名前'].unique():
    if name not in employee_main_base:
        emp_data = naisen_by_emp[naisen_by_emp['最終着信者_名前'] == name]
        main_base = emp_data.loc[emp_data['count'].idxmax(), '最終着信者_拠点']
        employee_main_base[name] = main_base

# 発信者としてのみ存在する場合も補完
for _, row in df_naisen.iterrows():
    if row['発信者_名前'] and row['発信者_拠点']:
        if row['発信者_名前'] not in employee_main_base:
            employee_main_base[row['発信者_名前']] = row['発信者_拠点']

for _, row in df_gaisen_hasshin.iterrows():
    if row['発信者_名前'] and row['発信者_拠点']:
        if row['発信者_名前'] not in employee_main_base:
            employee_main_base[row['発信者_名前']] = row['発信者_拠点']

employee_data = []
for name, base in employee_main_base.items():
    if name in EXCLUDE_NAMES:
        continue
    
    # 該当拠点での内線着信数（主要拠点ベース）
    naisen_count = len(df_naisen[(df_naisen['最終着信者_名前'] == name) & (df_naisen['最終着信者_拠点'] == base)])
    naisen_calls = df_naisen[(df_naisen['最終着信者_名前'] == name) & (df_naisen['最終着信者_拠点'] == base)]
    naisen_avg_time = naisen_calls['通話時間'].mean() if len(naisen_calls) > 0 else 0
    
    # 該当拠点での外線着信数（主要拠点ベース）
    gaisen_count = len(df_gaisen_chakushin[(df_gaisen_chakushin['最終着信者_名前'] == name) & 
                                            (df_gaisen_chakushin['最終着信者_拠点'] == base)])
    gaisen_calls = df_gaisen_chakushin[(df_gaisen_chakushin['最終着信者_名前'] == name) & 
                                        (df_gaisen_chakushin['最終着信者_拠点'] == base)]
    gaisen_avg_time = gaisen_calls['通話時間'].mean() if len(gaisen_calls) > 0 else 0
    
    juhatchu = '受発注' if name in JUHATCHU_LIST else None
    
    daily_counts = {}
    for date in date_range:
        d = date.date()
        naisen_daily = len(df_naisen[(df_naisen['最終着信者_名前'] == name) & 
                                     (df_naisen['最終着信者_拠点'] == base) & 
                                     (df_naisen['日付'] == d)])
        gaisen_daily = len(df_gaisen_chakushin[(df_gaisen_chakushin['最終着信者_名前'] == name) & 
                                               (df_gaisen_chakushin['最終着信者_拠点'] == base) & 
                                               (df_gaisen_chakushin['日付'] == d)])
        daily_counts[str(date.date())] = naisen_daily + gaisen_daily
    
    working_days = sum(1 for v in daily_counts.values() if v > 0)
    total_calls = naisen_count + gaisen_count
    avg_per_day = round(total_calls / working_days, 1) if working_days > 0 else 0
    
    row_data = {
        '名前': name, '拠点': base, '受発注': juhatchu,
        '内線': naisen_count, '通話時間／秒': round(naisen_avg_time, 1) if naisen_count > 0 else 0,
        '外線': gaisen_count, '外線_時間／秒': round(gaisen_avg_time, 1) if gaisen_count > 0 else 0,
    }
    row_data.update(daily_counts)
    row_data['稼働日'] = working_days
    row_data['内外線計'] = total_calls
    row_data['1日平均'] = avg_per_day
    employee_data.append(row_data)

base_order_map = {b: i for i, b in enumerate(BASE_ORDER)}
employee_data.sort(key=lambda x: (base_order_map.get(x['拠点'], 999), x['名前']))


# ============================================================
# シート3: 関数_拠点別
# ============================================================
print("シート3: 関数_拠点別を集計中...")

sheet3_data = []
for base in BASE_ORDER:
    gaisen = gaisen_chakuden.get(base, 0)
    employees_in_base = [e for e in employee_data if e['拠点'] == base]
    headcount = len(employees_in_base)
    per_person = gaisen / headcount if headcount > 0 else 0
    
    sheet3_data.append({
        '拠点名': base,
        '2025年12月から外線のみ': gaisen,
        '外線のみ': round(gaisen / 30, 1),
        '人員': headcount,
        '1人当たり／月': round(per_person, 1),
        '全体からの比率': gaisen / gaisen_chakuden.sum() if gaisen_chakuden.sum() > 0 else 0
    })

juhatchu_data = []
for base in JUHATCHU_BASE_ORDER:
    juhatchu_employees = [e for e in employee_data if e['拠点'] == base and e['受発注'] == '受発注']
    total_gaisen = sum(e['外線'] for e in juhatchu_employees)
    headcount = len(juhatchu_employees)
    per_person = total_gaisen / headcount if headcount > 0 else 0
    juhatchu_data.append({'拠点名': base, '受発注_外線': total_gaisen, '受発注_人員': headcount, '1人当たり／月': round(per_person, 1)})


# ============================================================
# シート4: 時短勤務
# ============================================================
print("シート4: 時短勤務を作成中...")

sheet4_data = []
for jitan in JITAN_DATA:
    emp = next((e for e in employee_data if e['名前'] == jitan['名前_key']), None)
    gaisen_actual = emp['外線'] if emp else 0
    keisu = jitan['係数']  # 見本の係数をそのまま使用
    sheet4_data.append({
        '氏名': jitan['氏名'], '部署': jitan['部署'], '勤務時間': jitan['勤務時間'],
        '係数': keisu, '外線(実績)': gaisen_actual, '外線(見込)': round(gaisen_actual * keisu, 1)
    })


# ============================================================
# シート5: 営業時間内集計（Final Base）
# ============================================================
print("シート5: 営業時間内集計を作成中...")

gaisen_business = df_gaisen_chakushin[df_gaisen_chakushin['営業時間内']]
gaisen_business_by_base = gaisen_business.groupby('最終着信者_拠点').size()

sheet5_data = []
for base in BASE_ORDER:
    gaisen = gaisen_business_by_base.get(base, 0)
    employees_in_base = [e for e in employee_data if e['拠点'] == base]
    headcount = len(employees_in_base)
    per_person = gaisen / headcount if headcount > 0 else 0
    total = gaisen_business_by_base.sum()
    ratio = gaisen / total if total > 0 else 0
    sheet5_data.append({
        '拠点名': base, '営業時間内_外線のみ': gaisen, '人員': headcount,
        '1人当たり／月': round(per_person, 1), '全体からの比率': round(ratio, 6)
    })


# ============================================================
# シート6: 時間内集計（Target Base）
# ============================================================
print("シート6: 時間内集計を作成中...")

naisen_business = df_naisen[df_naisen['営業時間内']]
naisen_nyuden_biz = naisen_business.groupby('着信者_拠点').size()
naisen_chakuden_biz = naisen_business.groupby('最終着信者_拠点').size()

gaisen_nyuden_biz = gaisen_business.groupby('着信者_拠点').size()
gaisen_chakuden_biz = gaisen_business.groupby('最終着信者_拠点').size()

sheet6_data = []
for base in BASE_ORDER:
    n_nyuden = naisen_nyuden_biz.get(base, 0)
    n_chakuden = naisen_chakuden_biz.get(base, 0)
    n_ratio = n_chakuden / n_nyuden if n_nyuden > 0 else 0
    
    g_nyuden = gaisen_nyuden_biz.get(base, 0)
    g_chakuden = gaisen_chakuden_biz.get(base, 0)
    g_ratio = g_chakuden / g_nyuden if g_nyuden > 0 else 0
    
    sheet6_data.append({
        '拠点': base,
        '内線_入電': n_nyuden, '内線_着電': n_chakuden, '内線_応答率': round(n_ratio, 6) if n_nyuden > 0 else 0,
        '外線_入電': g_nyuden, '外線_着電': g_chakuden, '外線_応答率': round(g_ratio, 6) if g_nyuden > 0 else 0
    })

total_n_nyuden = sum(r['内線_入電'] for r in sheet6_data)
total_n_chakuden = sum(r['内線_着電'] for r in sheet6_data)
total_g_nyuden = sum(r['外線_入電'] for r in sheet6_data)
total_g_chakuden = sum(r['外線_着電'] for r in sheet6_data)

sheet6_data.append({
    '拠点': '合計',
    '内線_入電': total_n_nyuden, '内線_着電': total_n_chakuden,
    '内線_応答率': round(total_n_chakuden / total_n_nyuden, 6) if total_n_nyuden > 0 else 0,
    '外線_入電': total_g_nyuden, '外線_着電': total_g_chakuden,
    '外線_応答率': round(total_g_chakuden / total_g_nyuden, 6) if total_g_nyuden > 0 else 0
})

additional_biz = []
for group_name, bases in REGION_GROUPS.items():
    nyuden = sum(gaisen_nyuden_biz.get(b, 0) for b in bases)
    chakuden = sum(gaisen_chakuden_biz.get(b, 0) for b in bases)
    ratio = chakuden / nyuden if nyuden > 0 else 0
    additional_biz.append({'グループ': group_name, '入った数': nyuden, 'とった数': chakuden, '％': ratio})

total_biz_nyuden = sum(r['入った数'] for r in additional_biz)
total_biz_chakuden = sum(r['とった数'] for r in additional_biz)
additional_biz.append({'グループ': '合計', '入った数': total_biz_nyuden, 'とった数': total_biz_chakuden,
                       '％': total_biz_chakuden / total_biz_nyuden if total_biz_nyuden > 0 else 0})


# ============================================================
# Excelファイル出力
# ============================================================
print("Excelファイルを出力中...")

wb = Workbook()

# シート1: 着信件数
ws1 = wb.active
ws1.title = '1.着信件数'

ws1['A1'] = '拠点'
ws1['B1'] = '内線'
ws1['D1'] = '外線'
ws1['F1'] = '他拠点へ転送'
ws1['G1'] = '他拠点から転送'
ws1['B2'] = '入電'
ws1['C2'] = '着電'
ws1['D2'] = '入電'
ws1['E2'] = '着電'

for i, row in enumerate(sheet1_data):
    ws1[f'A{i+3}'] = row['拠点']
    ws1[f'B{i+3}'] = row['内線_入電']
    ws1[f'C{i+3}'] = row['内線_着電']
    ws1[f'D{i+3}'] = row['外線_入電']
    ws1[f'E{i+3}'] = row['外線_着電']
    ws1[f'F{i+3}'] = row['他拠点へ転送']
    ws1[f'G{i+3}'] = row['他拠点から転送']

ws1['J26'] = '追加集計(24H)'
ws1['J27'] = '電話が入った数'
ws1['K27'] = 'とった数'
ws1['L27'] = '％'
for i, row in enumerate(additional_24h):
    ws1[f'J{i+28}'] = row['グループ']
    ws1[f'K{i+28}'] = row['電話が入った数']
    ws1[f'L{i+28}'] = row['とった数']
    ws1[f'M{i+28}'] = row['％']

# シート2: 従業員別
ws2 = wb.create_sheet('2.従業員別')
date_columns = [str(d.date()) for d in date_range]
columns = ['名前', '拠点', '受発注', '内線', '通話時間／秒', '外線', '外線_時間／秒'] + date_columns + ['稼働日', '内外線計', '1日平均']

for col_idx, col_name in enumerate(columns, 1):
    ws2.cell(row=1, column=col_idx, value=col_name)

for row_idx, emp in enumerate(employee_data, 2):
    ws2.cell(row=row_idx, column=1, value=emp['名前'])
    ws2.cell(row=row_idx, column=2, value=emp['拠点'])
    ws2.cell(row=row_idx, column=3, value=emp['受発注'])
    ws2.cell(row=row_idx, column=4, value=emp['内線'])
    ws2.cell(row=row_idx, column=5, value=emp['通話時間／秒'])
    ws2.cell(row=row_idx, column=6, value=emp['外線'])
    ws2.cell(row=row_idx, column=7, value=emp['外線_時間／秒'])
    for date_idx, date_col in enumerate(date_columns):
        ws2.cell(row=row_idx, column=8+date_idx, value=emp.get(date_col, 0))
    ws2.cell(row=row_idx, column=8+len(date_columns), value=emp['稼働日'])
    ws2.cell(row=row_idx, column=9+len(date_columns), value=emp['内外線計'])
    ws2.cell(row=row_idx, column=10+len(date_columns), value=emp['1日平均'])

# シート3: 関数_拠点別
ws3 = wb.create_sheet('3.関数_拠点別')
headers3 = ['拠点名', '2025年12月から外線のみ', '外線のみ', '人員', '1人当たり／月', '全体からの比率']
for col_idx, h in enumerate(headers3, 1):
    ws3.cell(row=1, column=col_idx, value=h)

for row_idx, row in enumerate(sheet3_data, 2):
    ws3.cell(row=row_idx, column=1, value=row['拠点名'])
    ws3.cell(row=row_idx, column=2, value=row['2025年12月から外線のみ'])
    ws3.cell(row=row_idx, column=3, value=row['外線のみ'])
    ws3.cell(row=row_idx, column=4, value=row['人員'])
    ws3.cell(row=row_idx, column=5, value=row['1人当たり／月'])
    ws3.cell(row=row_idx, column=6, value=row['全体からの比率'])

ws3['I1'] = '拠点名'
ws3['J1'] = '受発注_外線'
ws3['K1'] = '受発注_人員'
ws3['L1'] = '1人当たり／月'
for row_idx, row in enumerate(juhatchu_data, 2):
    ws3.cell(row=row_idx, column=9, value=row['拠点名'])
    ws3.cell(row=row_idx, column=10, value=row['受発注_外線'])
    ws3.cell(row=row_idx, column=11, value=row['受発注_人員'])
    ws3.cell(row=row_idx, column=12, value=row['1人当たり／月'])

# シート4: 時短勤務
ws4 = wb.create_sheet('4.時短勤務')
headers4 = ['氏名', '部署', '勤務時間', '係数', '外線(実績)', '外線(見込)']
for col_idx, h in enumerate(headers4, 1):
    ws4.cell(row=1, column=col_idx, value=h)
for row_idx, row in enumerate(sheet4_data, 2):
    ws4.cell(row=row_idx, column=1, value=row['氏名'])
    ws4.cell(row=row_idx, column=2, value=row['部署'])
    ws4.cell(row=row_idx, column=3, value=row['勤務時間'])
    ws4.cell(row=row_idx, column=4, value=row['係数'])
    ws4.cell(row=row_idx, column=5, value=row['外線(実績)'])
    ws4.cell(row=row_idx, column=6, value=row['外線(見込)'])

# シート5: 営業時間内集計
ws5 = wb.create_sheet('5.営業時間内集計')
ws5['A1'] = "※集計基準：『誰が取ったか（Final Base）』でカウント（例：東京の人が流山宛ての外線を取ったら『東京』の実績になります）"
headers5 = ['拠点名', '営業時間内_外線のみ', '人員', '1人当たり／月', '全体からの比率']
for col_idx, h in enumerate(headers5, 1):
    ws5.cell(row=2, column=col_idx, value=h)
for row_idx, row in enumerate(sheet5_data, 3):
    ws5.cell(row=row_idx, column=1, value=row['拠点名'])
    ws5.cell(row=row_idx, column=2, value=row['営業時間内_外線のみ'])
    ws5.cell(row=row_idx, column=3, value=row['人員'])
    ws5.cell(row=row_idx, column=4, value=row['1人当たり／月'])
    ws5.cell(row=row_idx, column=5, value=row['全体からの比率'])

# シート6: 時間内集計
ws6 = wb.create_sheet('6.時間内集計')
ws6['A1'] = "※集計基準：『どこ宛てか（Target Base）』でカウント（例：東京の人が流山宛ての外線を取っても、流山に着信したので『流山』のカウントになります）"
ws6['A3'] = '拠点'
ws6['B3'] = '内線'
ws6['E3'] = '外線'
ws6['B5'] = '入電'
ws6['C5'] = '着電'
ws6['D5'] = '応答率'
ws6['E5'] = '入電'
ws6['F5'] = '着電'
ws6['G5'] = '応答率'

for row_idx, row in enumerate(sheet6_data, 6):
    ws6.cell(row=row_idx, column=1, value=row['拠点'])
    ws6.cell(row=row_idx, column=2, value=row['内線_入電'])
    ws6.cell(row=row_idx, column=3, value=row['内線_着電'])
    ws6.cell(row=row_idx, column=4, value=row['内線_応答率'])
    ws6.cell(row=row_idx, column=5, value=row['外線_入電'])
    ws6.cell(row=row_idx, column=6, value=row['外線_着電'])
    ws6.cell(row=row_idx, column=7, value=row['外線_応答率'])

ws6['M5'] = '表計(入電)'
ws6['N5'] = '全入電(N1)'
ws6['O5'] = '全不在(O1)'
ws6['P5'] = '全着電(P1)'
ws6['Q5'] = '全応答率(Q1)'

total_all_nyuden = len(df_gaisen_chakushin[df_gaisen_chakushin['営業時間内']])
total_all_fuzai = len(df_gaisen_chakushin[(df_gaisen_chakushin['営業時間内']) & (df_gaisen_chakushin['最終着信者'] == '不在')])
total_all_chakuden = total_all_nyuden - total_all_fuzai
total_all_ratio = total_all_chakuden / total_all_nyuden if total_all_nyuden > 0 else 0

ws6['M6'] = total_all_nyuden
ws6['N6'] = total_all_nyuden
ws6['O6'] = total_all_fuzai
ws6['P6'] = total_all_chakuden
ws6['Q6'] = total_all_ratio

ws6['M9'] = '追加集計(時間内)'
ws6['N9'] = '入った数'
ws6['O9'] = 'とった数'
ws6['P9'] = '％'
for i, row in enumerate(additional_biz):
    ws6.cell(row=10+i, column=13, value=row['グループ'])
    ws6.cell(row=10+i, column=14, value=row['入った数'])
    ws6.cell(row=10+i, column=15, value=row['とった数'])
    ws6.cell(row=10+i, column=16, value=row['％'])

wb.save(OUTPUT_FILE)
print(f"\n完了！出力ファイル: {OUTPUT_FILE}")
print(f"シート1: 着信件数 - {len(sheet1_data)}行")
print(f"シート2: 従業員別 - {len(employee_data)}人")
print(f"シート3: 関数_拠点別 - {len(sheet3_data)}拠点")
print(f"シート4: 時短勤務 - {len(sheet4_data)}人")
print(f"シート5: 営業時間内集計 - {len(sheet5_data)}拠点")
print(f"シート6: 時間内集計 - {len(sheet6_data)}行")
