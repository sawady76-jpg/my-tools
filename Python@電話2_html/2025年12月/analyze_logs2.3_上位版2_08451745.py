import pandas as pd
import glob
import re
import os
import sys
import datetime
import warnings

# 警告を無視
warnings.simplefilter('ignore')

# ★ここがポイント：スクリプトのある場所をカレントディレクトリに設定
try:
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
except:
    pass

# ==========================================
# 1. 設定エリア
# ==========================================
BASE_ORDER = [
    '東京', '横浜', '埼玉', '滋賀', '大阪', 
    '千葉', '福岡', '岡山', '名古屋', '仙台', '流山'
]
EXCLUDE_KEYWORDS = ['不在', '未応答', '応答なし', '放棄', '留守電']
BIZ_START = datetime.time(8, 45, 0)
BIZ_END = datetime.time(17, 45, 0)
TRANS_FROM = {
    '東京': '流山', '横浜': '多摩', '埼玉': '千葉・北関東',
    '滋賀': '金沢', '福岡': '広島・熊本', '岡山': '大阪・静岡', '仙台': '札幌'
}
TRANS_TO = {'千葉': '埼玉', '大阪': '岡山'}

SUMMARY_GROUPS = [
    {'name': '東京+流山', 'members': ['東京', '流山']},
    {'name': '埼玉+千葉+北関東', 'members': ['埼玉', '千葉', '北関東']},
    {'name': '横浜+多摩', 'members': ['横浜', '多摩']},
    {'name': '仙台+札幌', 'members': ['仙台', '札幌']},
    {'name': '滋賀+金沢', 'members': ['滋賀', '金沢']},
    {'name': '名古屋', 'members': ['名古屋']},
    {'name': '福岡+熊本+広島', 'members': ['福岡', '熊本', '広島']},
    {'name': '岡山+大阪+静岡', 'members': ['岡山', '大阪', '静岡']}
]

# 時短勤務者の名字→フルネーム変換マッピング
JITAN_FULLNAME_MAP = {
    "玉腰": "玉腰　千恵",
    "石川": "石川　恵理",
    "矢島": "矢島　静音",
    "河原": "河原　由布子"
}

# ★時短勤務者4名の固定データ（氏名, 部署, 勤務時間）
JITAN_MEMBERS = [
    {"氏名": "玉腰　千恵", "部署": "名古屋営業所(業務)", "勤務時間": 5.75},
    {"氏名": "石川　恵理", "部署": "埼玉支店(業務)", "勤務時間": 6.25},
    {"氏名": "矢島　静音", "部署": "埼玉支店(業務)", "勤務時間": 6},
    {"氏名": "河原　由布子", "部署": "岡山営業所(業務)", "勤務時間": 6},
]

JUHATCHU_MEMBERS_RAW = [
    "【流山】鴻池千紗都", "【横浜】奥秋素子", "【横浜】山﨑啓右", "【横浜】青木観奈",
    "【横浜】川崎瞳", "【横浜】中山満里奈", "【横浜】萩原直美", "【横浜】舞智江",
    "【岡山】伊藤優", "【岡山】河原由布子", "【岡山】岸本麻衣子", "【岡山】岩佐寛子",
    "【岡山】武政彩果", "【岡山】矢野夏絵", "【埼玉】竹内梓", "【埼玉】吉田夏子",
    "【埼玉】金子友希", "【埼玉】細田昌子", "【埼玉】石川恵理", "【埼玉】池田千恵子",
    "【埼玉】豊田里花", "【埼玉】矢島静音", "【滋賀】松井治美", "【滋賀】梅原薫",
    "【滋賀】筆坂雪子", "【滋賀】北出智子", "【仙台】阿部かおり", "【仙台】真壁彰子",
    "【仙台】仁藤佳美", "【仙台】齋藤詩織", "【千葉】小島佳菜", "【大阪】松下愛",
    "【東京】高澤早紀", "【東京】佐々木美咲", "【東京】坂田智世", "【東京】西垣彩",
    "【東京】北林友希", "【東京】林まど佳", "【福岡】奥薗美和", "【福岡】山崎芹奈",
    "【福岡】重松宝成", "【福岡】松添愛", "【福岡】松尾明日香", "【名古屋】稲垣みちる",
    "【名古屋】玉腰千恵", "【名古屋】水島奈美", "【名古屋】大渕温子"
]

# ==========================================
# 2. 関数定義
# ==========================================

def extract_base_name(text):
    if not isinstance(text, str): return None, None
    match = re.search(r'【(.*?)】(.*)', text)
    if match: return match.group(1), match.group(2).strip()
    return None, text

def is_valid_answer(name):
    if not isinstance(name, str): return False
    for kw in EXCLUDE_KEYWORDS:
        if kw in name: return False
    return True

def my_round(x):
    try:
        if pd.isna(x): return ""
        val = float(x)
        return int(val * 10 + 0.5) / 10.0
    except: return x

def smart_read_excel(file, sheet_name):
    try:
        preview = pd.read_excel(file, sheet_name=sheet_name, header=None, nrows=20)
        target_row = 0
        for i, row in preview.iterrows():
            row_str = row.astype(str).values
            if any('着信者' in s for s in row_str) or any('時刻' in s for s in row_str):
                target_row = i
                break
        return pd.read_excel(file, sheet_name=sheet_name, header=target_row)
    except:
        return pd.DataFrame()

def smart_read_jitan(file, sheet_name):
    """時短勤務ファイル用のヘッダー自動検出"""
    try:
        preview = pd.read_excel(file, sheet_name=sheet_name, header=None, nrows=20)
        target_row = 0
        for i, row in preview.iterrows():
            # 行内の値を文字列リストに変換
            row_str = [str(x).strip() for x in row.values]
            # 「氏名」または「名前」があり、かつ「勤務時間」が含まれる行を探す
            if ('氏名' in row_str or '名前' in row_str) and '勤務時間' in row_str:
                target_row = i
                break
        return pd.read_excel(file, sheet_name=sheet_name, header=target_row)
    except:
        return pd.DataFrame()

def find_and_load(target_keywords, exclude_keywords=[]):
    candidates = []
    for file in glob.glob("*.xlsx"):
        if "集計結果" in file: continue
        try:
            xls = pd.ExcelFile(file)
            for sheet in xls.sheet_names:
                if any(k in sheet for k in target_keywords):
                    if not any(ek in sheet for ek in exclude_keywords):
                        candidates.append({'file': file, 'sheet': sheet, 'type': 'xlsx'})
        except: pass

    for file in glob.glob("*.csv"):
        if "集計結果" in file: continue
        if any(k in file for k in target_keywords):
            if not any(ek in file for ek in exclude_keywords):
                candidates.append({'file': file, 'sheet': None, 'type': 'csv'})

    if not candidates:
        return pd.DataFrame()

    candidates.sort(key=lambda x: (0 if "統合版" in x['file'] else 1))
    
    target = candidates[0]
    print(f"  -> 読み込み: {os.path.basename(target['file'])} (Sheet: {target['sheet'] if target['sheet'] else 'CSV'})")
    
    if target['type'] == 'xlsx':
        # ★時短ファイルは専用の読み込み関数を使う
        if any('時短' in k for k in target_keywords):
            return smart_read_jitan(target['file'], target['sheet'])
        else:
            return smart_read_excel(target['file'], target['sheet'])
    else:
        try: return pd.read_csv(target['file'], encoding='utf-8')
        except: return pd.read_csv(target['file'], encoding='cp932')

# ★★★ デザイン装飾関数（修正版） ★★★
def decorate_excel(filename):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxlがインストールされていないため、装飾をスキップします。")
        return

    wb = openpyxl.load_workbook(filename)
    
    # --- スタイル定義 ---
    FONT_NAME = "BIZ UDゴシック"
    
    NAVY = "1F497D"
    PALE_BLUE = "DCE6F1"
    LIGHT_YELLOW = "FFF2CC"
    WHITE = "FFFFFF"
    GRAY_BORDER = "BFBFBF"
    
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid") 
    header_font = Font(bold=True, color=WHITE, name=FONT_NAME)
    
    sub_header_fill = PatternFill(start_color=PALE_BLUE, end_color=PALE_BLUE, fill_type="solid") 
    sub_header_font = Font(bold=True, name=FONT_NAME)
    
    total_row_fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid") 
    white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")
    
    thin_border = Side(style='thin', color=GRAY_BORDER)
    border_all = Border(left=thin_border, right=thin_border, top=thin_border, bottom=thin_border)
    
    base_font = Font(name=FONT_NAME)
    bold_font = Font(bold=True, name=FONT_NAME)
    note_font = Font(name=FONT_NAME, color="FF0000", size=10, bold=True)

    def set_cell_style(cell, style_type="data", number=None):
        cell.border = border_all
        cell.font = base_font
        
        # 数値書式設定
        if number is not None:
            if isinstance(number, (int, float)):
                # 率の判定
                str_val = str(cell.value)
                if "率" in str_val or "％" in str_val:
                     cell.number_format = '0.0%' 
                elif isinstance(number, float):
                     cell.number_format = '#,##0.0'
                else:
                     cell.number_format = '#,##0'
        
        if style_type == "header":
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        elif style_type == "sub_header":
            cell.fill = sub_header_fill
            cell.font = sub_header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        elif style_type == "data":
            cell.fill = white_fill
        elif style_type == "total":
            cell.fill = total_row_fill
            cell.font = bold_font

    def clean_range(ws, min_r, min_c, max_r, max_c):
        for r in range(min_r, max_r + 1):
            for c in range(min_c, max_c + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill(fill_type=None)
                cell.border = None
                cell.value = ""

    def auto_fit(ws):
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            adjusted_width = (max_length + 2) * 1.1
            ws.column_dimensions[get_column_letter(column[0].column)].width = min(adjusted_width, 50)

    # --- メイン処理 ---
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.sheet_view.showGridLines = False
        
        for row in ws.iter_rows():
            for cell in row:
                cell.font = base_font

        auto_fit(ws)
        max_r = ws.max_row

        if sheet_name == '1.着信件数':
            # --- メイン表 (A1:G[max]) ---
            for r in range(1, 3):
                for c in range(1, 8):
                    set_cell_style(ws.cell(row=r, column=c), "header")
            
            ws['A1'] = "拠点"
            ws.merge_cells('A1:A2')
            ws.merge_cells('B1:C1') # 内線
            ws.merge_cells('D1:E1') # 外線
            ws.merge_cells('F1:F2') # 転送
            ws.merge_cells('G1:G2') # 転送
            
            # ★データ開始行を動的に検出（空行をスキップ）
            first_data_row = 3
            for r in range(3, 10):
                if ws.cell(row=r, column=1).value is not None:
                    first_data_row = r
                    break
            
            # データ終了行を検出
            real_max_r_main = 0
            for r in range(first_data_row, max_r + 20):
                if ws.cell(row=r, column=1).value is None and ws.cell(row=r, column=2).value is None:
                    real_max_r_main = r - 1
                    break
            
            if real_max_r_main < first_data_row: real_max_r_main = max_r 

            # ★データ行にスタイル適用（動的開始行から）
            for r in range(first_data_row, real_max_r_main + 1):
                style = "total" if ws.cell(row=r, column=1).value == "合計" else "data"
                for c in range(1, 8):
                    set_cell_style(ws.cell(row=r, column=c), style, ws.cell(row=r, column=c).value)

            # 余白クリーニング (H列)
            clean_range(ws, 1, 8, max_r + 5, 9)
            ws.column_dimensions['H'].width = 2 
            
            # ★追加集計エリアの列幅を設定（レイアウト崩れ防止）
            ws.column_dimensions['I'].width = 2   # 余白
            ws.column_dimensions['J'].width = 18  # 追加集計(24H) / グループ名
            ws.column_dimensions['K'].width = 14  # 電話が入った数
            ws.column_dimensions['L'].width = 10  # とった数
            ws.column_dimensions['M'].width = 10  # ％

            # --- 追加集計 (J列エリア) ---
            # グループ集計 (24H)
            for row in ws.iter_rows(min_col=10, max_col=10):
                for cell in row:
                    if cell.value == "追加集計(24H)":
                        start_r = cell.row
                        end_r = start_r + len(SUMMARY_GROUPS) + 2
                        
                        # ヘッダー行（2行分）
                        for r in range(start_r, start_r + 2):
                            for c in range(10, 14):
                                set_cell_style(ws.cell(row=r, column=c), "sub_header")
                        
                        # データ行
                        for r in range(start_r + 2, end_r + 1):
                            style = "total" if r == end_r else "data"
                            for c in range(10, 14):
                                val = ws.cell(row=r, column=c).value
                                set_cell_style(ws.cell(row=r, column=c), style, val)
                                if c == 13 and isinstance(val, (int, float)):
                                    ws.cell(row=r, column=c).number_format = '0.0%'
                        break

        elif sheet_name == '5.営業時間内集計':
            ws.insert_rows(1)
            ws['A1'] = "※集計基準：『誰が取ったか（Final Base）』でカウント（例：東京の人が流山宛ての外線を取ったら『東京』の実績になります）"
            ws['A1'].font = note_font
            
            for c in range(1, ws.max_column + 1):
                set_cell_style(ws.cell(row=2, column=c), "header")
            
            for r in range(3, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    val = ws.cell(row=r, column=c).value
                    if "比率" in str(ws.cell(row=2, column=c).value):
                         if isinstance(val, (int, float)):
                             ws.cell(row=r, column=c).number_format = '0.0%'
                    set_cell_style(ws.cell(row=r, column=c), "data", val)

        elif sheet_name == '6.時間内集計':
            ws.insert_rows(1)
            # 1行挿入したため、Pandasが出力したデータ（元4行目開始）は「5行目」に移動しています。
            
            ws['A1'] = "※集計基準：『どこ宛てか（Target Base）』でカウント（例：東京の人が流山宛ての外線を取っても、流山に着信したので『流山』のカウントになります）"
            ws['A1'].font = note_font
            
            max_r = ws.max_row 
            
            # --- メイン表 (A5起点) ---
            for r in range(5, 7): # ヘッダーは5,6行目
                for c in range(1, 8):
                    set_cell_style(ws.cell(row=r, column=c), "header")
            
            ws['A5'] = "拠点"
            ws.merge_cells('A5:A6') 
            ws.merge_cells('B5:D5') # 内線
            ws.merge_cells('E5:G5') # 外線
            
            # データ部分は7行目から
            for r in range(7, max_r + 1):
                # 合計行の判定（A列が"合計"かどうか）
                style = "total" if ws.cell(row=r, column=1).value == "合計" else "data"
                for c in range(1, 8):
                    val = ws.cell(row=r, column=c).value
                    set_cell_style(ws.cell(row=r, column=c), style, val)
                    if c in [4, 7] and isinstance(val, (int, float)):
                        ws.cell(row=r, column=c).number_format = '0.0%'

            # 余白クリーニング (H-L列)
            clean_range(ws, 1, 8, max_r + 5, 12)
            
            # --- 上部集計 (M2:Q3) ---
            for c in range(13, 18): # M-Q
                set_cell_style(ws.cell(row=2, column=c), "sub_header")
                val = ws.cell(row=3, column=c).value
                set_cell_style(ws.cell(row=3, column=c), "data", val)
                ws.cell(row=3, column=c).font = Font(bold=True, name=FONT_NAME)
                if c == 17 and isinstance(val, (int, float)):
                    ws.cell(row=3, column=c).number_format = '0.0%'

            # --- グループ集計 (M6起点) ---
            start_r = 6 
            end_r = start_r + len(SUMMARY_GROUPS) + 2 + 1
            
            for r in range(start_r, start_r + 2):
                for c in range(13, 17):
                    set_cell_style(ws.cell(row=r, column=c), "sub_header")
            
            for r in range(start_r + 2, end_r + 1):
                style = "total" if r == end_r else "data"
                for c in range(13, 17):
                    val = ws.cell(row=r, column=c).value
                    set_cell_style(ws.cell(row=r, column=c), style, val)
                    if c == 16 and isinstance(val, (int, float)):
                        ws.cell(row=r, column=c).number_format = '0.0%'

        else:
            for c in range(1, ws.max_column + 1):
                set_cell_style(ws.cell(row=1, column=c), "header")
            for r in range(2, ws.max_row + 1):
                for c in range(1, ws.max_column + 1):
                    val = ws.cell(row=r, column=c).value
                    set_cell_style(ws.cell(row=r, column=c), "data", val)

    wb.save(filename)
    print(" -> Excelファイルの装飾・数値書式設定が完了しました。")

# ==========================================
# 3. メイン処理
# ==========================================
def main():
    print(f"作業フォルダ: {os.getcwd()}")
    print("データを探しています...")

    df_int = find_and_load(['内線通話', '内線'], exclude_keywords=['発信'])
    df_ext = find_and_load(['外線着信', '外線'], exclude_keywords=['発信'])
    # df_short = find_and_load(['時短勤務', '時短'])  # ★外部ファイル読み込みを無効化（固定データを使用）

    print(f"\nデータ件数: 内線={len(df_int)}件, 外線={len(df_ext)}件, 時短=4名(固定)")

    if df_int.empty and df_ext.empty:
        print("\n【エラー】データが見つかりません。")
        return

    data_year_month = None
    for df in [df_int, df_ext]:
        if df.empty: continue
        df.columns = [str(c).strip() for c in df.columns]
        if '着信者' in df.columns:
            df['target_base'], _ = zip(*df['着信者'].apply(extract_base_name))
        else: df['target_base'] = None
        if '最終着信者' in df.columns:
            df['final_base'], df['final_name'] = zip(*df['最終着信者'].apply(extract_base_name))
        else:
            df['final_base'] = None
            df['final_name'] = None
        if '時刻' in df.columns:
            df['dt'] = pd.to_datetime(df['時刻'].astype(str), errors='coerce')
            df['day'] = df['dt'].dt.day
            df['time'] = df['dt'].dt.time
            df['date'] = df['dt'].dt.date
            if data_year_month is None and not df['dt'].dropna().empty:
                first_date = df['dt'].dropna().iloc[0]
                data_year_month = (first_date.year, first_date.month)

    all_dates = set()
    if not df_int.empty: all_dates.update(df_int['date'].dropna().unique())
    if not df_ext.empty: all_dates.update(df_ext['date'].dropna().unique())
    total_operating_days = len(all_dates) if all_dates else 1
    
    # --- Sheet 1 Data ---
    if not df_int.empty and 'final_name' in df_int.columns:
        df_int_valid = df_int[df_int['final_name'].apply(is_valid_answer)].copy()
    else: df_int_valid = pd.DataFrame()

    if not df_ext.empty and 'final_name' in df_ext.columns:
        df_ext_valid = df_ext[df_ext['final_name'].apply(is_valid_answer)].copy()
    else: df_ext_valid = pd.DataFrame()

    all_bases = set(BASE_ORDER)
    if not df_int.empty and 'target_base' in df_int.columns: all_bases |= set(df_int['target_base'].dropna())
    if not df_ext.empty and 'target_base' in df_ext.columns: all_bases |= set(df_ext['target_base'].dropna())
    sorted_bases = sorted(list(all_bases), key=lambda x: BASE_ORDER.index(x) if x in BASE_ORDER else 999)

    s1 = pd.DataFrame(index=sorted_bases)
    s1.index.name = '拠点'
    
    def get_counts(df, col):
        if df.empty or col not in df.columns: return {}
        return df[col].value_counts()

    s1_data = {
        ('内線', '入電'): s1.index.map(get_counts(df_int, 'target_base')).fillna(0).astype(int),
        ('内線', '着電'): s1.index.map(get_counts(df_int_valid, 'final_base')).fillna(0).astype(int),
        ('外線', '入電'): s1.index.map(get_counts(df_ext, 'target_base')).fillna(0).astype(int),
        ('外線', '着電'): s1.index.map(get_counts(df_ext_valid, 'final_base')).fillna(0).astype(int),
        ('他拠点へ転送', ' '): s1.index.map(TRANS_TO).fillna(''),
        ('他拠点から転送', ' '): s1.index.map(TRANS_FROM).fillna('')
    }
    s1 = pd.DataFrame(s1_data, index=s1.index)

    s1_total = s1.sum(numeric_only=True)
    s1.loc['合計'] = s1_total
    s1.loc['合計', ('他拠点へ転送', ' ')] = ''
    s1.loc['合計', ('他拠点から転送', ' ')] = ''

    summary_rows = [['追加集計(24H)', '電話が入った数', 'とった数', '％']]
    grp_total_in = 0
    grp_total_ans = 0
    for grp in SUMMARY_GROUPS:
        members = [m for m in grp['members'] if m in s1.index and m != '合計']
        if members:
            sum_in = s1.loc[members, ('外線', '入電')].sum()
            sum_ans = s1.loc[members, ('外線', '着電')].sum()
            rate = sum_ans / sum_in if sum_in > 0 else 0.0 
            summary_rows.append([grp['name'], sum_in, sum_ans, rate])
            grp_total_in += sum_in
            grp_total_ans += sum_ans
        else:
            summary_rows.append([grp['name'], 0, 0, 0.0])
    grp_total_rate = grp_total_ans / grp_total_in if grp_total_in > 0 else 0.0
    summary_rows.append(['合計', grp_total_in, grp_total_ans, grp_total_rate])
    s1_summary = pd.DataFrame(summary_rows)

    # --- Sheet 2 ---
    cols = ['final_name', 'final_base', '通話時間', 'day', 'date']
    s2 = pd.DataFrame()

    if (not df_int_valid.empty) or (not df_ext_valid.empty):
        if not df_int_valid.empty and set(cols).issubset(df_int_valid.columns):
            grp_int = df_int_valid.groupby(['final_name', 'final_base'])
            s2_int = grp_int.agg(内線件数=('通話時間', 'size'), 内線合計=('通話時間', 'sum'))
        else: s2_int = pd.DataFrame(columns=['内線件数', '内線合計'])

        if not df_ext_valid.empty and set(cols).issubset(df_ext_valid.columns):
            grp_ext = df_ext_valid.groupby(['final_name', 'final_base'])
            s2_ext = grp_ext.agg(外線件数=('通話時間', 'size'), 外線合計=('通話時間', 'sum'))
        else: s2_ext = pd.DataFrame(columns=['外線件数', '外線合計'])

        s2 = s2_int.join(s2_ext, how='outer').fillna(0)
        
        juhatchu_set = set()
        for raw_name in JUHATCHU_MEMBERS_RAW:
            b, n = extract_base_name(raw_name)
            if b and n:
                juhatchu_set.add((b, n))
        
        s2['受発注'] = s2.index.map(lambda x: "受発注" if (x[1], x[0]) in juhatchu_set else "")

        s2['内線'] = s2['内線件数']
        s2['通話時間／秒'] = (s2['内線合計'] / s2['内線件数'].replace(0, 1)).apply(my_round)
        s2['外線'] = s2['外線件数']
        s2['外線_時間／秒'] = (s2['外線合計'] / s2['外線件数'].replace(0, 1)).apply(my_round)
        
        df_all_valid = pd.concat([df_int_valid[cols] if not df_int_valid.empty else pd.DataFrame(), df_ext_valid[cols] if not df_ext_valid.empty else pd.DataFrame()])
        
        if not df_all_valid.empty:
            if data_year_month:
                y, m = data_year_month
                daily = df_all_valid.pivot_table(index=['final_name', 'final_base'], columns='day', aggfunc='size', fill_value=0)
                date_cols = {}
                for d in range(1, 32):
                    try:
                        date_str = datetime.date(y, m, d).strftime('%Y-%m-%d')
                        if d not in daily.columns: daily[d] = 0
                        date_cols[d] = date_str
                    except ValueError: pass
                daily = daily.rename(columns=date_cols)
                daily = daily[sorted(list(date_cols.values()))]
                daily['稼働日'] = (daily > 0).sum(axis=1)
                s2 = s2.join(daily).reset_index()
                s2['内外線計'] = s2['内線'] + s2['外線']
                s2['1日平均'] = s2.apply(lambda r: my_round(r['内外線計'] / r['稼働日']) if r['稼働日'] > 0 else 0, axis=1)
                s2 = s2.rename(columns={'final_name': '名前', 'final_base': '拠点'})
                s2['base_rank'] = s2['拠点'].apply(lambda x: BASE_ORDER.index(x) if x in BASE_ORDER else 999)
                s2 = s2.sort_values(['base_rank', '名前']).drop(columns='base_rank')
                final_cols = ['名前', '拠点', '受発注', '内線', '通話時間／秒', '外線', '外線_時間／秒'] + sorted(list(date_cols.values())) + ['稼働日', '内外線計', '1日平均']
                s2 = s2[[c for c in final_cols if c in s2.columns]]

    # --- Sheet 3 ---
    if ('外線', '着電') in s1.columns:
        s3 = s1[[('外線', '着電')]].reset_index().copy()
        s3.columns = ['拠点名', '外線(実績)']
    else: s3 = pd.DataFrame(columns=['拠点名', '外線(実績)'])
    
    s3 = s3[s3['拠点名'] != '合計'].copy()
    col_name_total = f"{data_year_month[0]}年{data_year_month[1]}月から外線のみ" if data_year_month else "期間計"
    s3 = s3.rename(columns={'外線(実績)': col_name_total})
    s3['外線のみ'] = s3[col_name_total].apply(lambda x: my_round(x / total_operating_days) if total_operating_days > 0 else 0)

    if not s2.empty:
        personnel = s2.groupby('拠点')['名前'].nunique()
        s3['人員'] = s3['拠点名'].map(personnel).fillna(0).astype(int)
    else: s3['人員'] = 0

    s3['1人当たり／月'] = s3.apply(lambda r: my_round(r[col_name_total] / r['人員']) if r['人員'] > 0 else 0.0, axis=1)
    total_s3 = s3[col_name_total].sum()
    s3['全体からの比率'] = s3[col_name_total].apply(lambda x: x / total_s3 if total_s3 > 0 else 0.0)
    s3 = s3[['拠点名', col_name_total, '外線のみ', '人員', '1人当たり／月', '全体からの比率']]

    # ★受発注グループ集計を作成
    s3_juhatchu = pd.DataFrame()
    if not s2.empty and '拠点' in s2.columns and '受発注' in s2.columns:
        # 受発注のある人のみフィルタ
        s2_juhatchu = s2[s2['受発注'] == '受発注'].copy()
        if not s2_juhatchu.empty:
            # 拠点別に外線件数を集計
            juhatchu_by_base = s2_juhatchu.groupby('拠点').agg(
                受発注_外線=('外線', 'sum'),
                受発注_人員=('名前', 'nunique')
            ).reset_index()
            juhatchu_by_base.columns = ['拠点名', '受発注_外線', '受発注_人員']
            s3_juhatchu = juhatchu_by_base

    # --- Sheet 4 (Jitan Processing) ---
    # ★時短勤務者データを固定設定から生成（外部ファイル不要）
    df_short = pd.DataFrame(JITAN_MEMBERS)
    
    # Helper: Normalize Name (remove spaces)
    def normalize_name(n):
        if not isinstance(n, str): return ""
        return re.sub(r'[\s　]+', '', n) # 全角半角スペース除去

    # Prepare lookup dictionary from s2
    if not s2.empty and '名前' in s2.columns:
        s2_temp = s2.copy()
        s2_temp['name_clean'] = s2_temp['名前'].apply(normalize_name)
        name_to_calls = s2_temp.set_index('name_clean')['外線'].to_dict()
    else:
        name_to_calls = {}

    # 係数計算（小数点第二位まで）
    def calc_coeff(hours):
        try:
            h = float(hours)
            if h <= 0: return 0
            return int((8.0 / h) * 100 + 0.5) / 100.0
        except: return 0
    
    df_short['係数'] = df_short['勤務時間'].apply(calc_coeff)
    
    # 外線実績を通話ログから取得
    def get_actual_calls(name):
        lookup_key = normalize_name(name)
        return name_to_calls.get(lookup_key, 0)
    
    df_short['外線(実績)'] = df_short['氏名'].apply(get_actual_calls)
    df_short['外線(実績)'] = pd.to_numeric(df_short['外線(実績)'], errors='coerce').fillna(0).astype(int)
    
    # 見込計算
    df_short['外線(見込)'] = df_short.apply(lambda r: my_round(r['外線(実績)'] * r['係数']), axis=1)
    
    # 列整理
    df_short = df_short[['氏名', '部署', '勤務時間', '係数', '外線(実績)', '外線(見込)']]
    
    # 数値丸め（係数は小数点第二位のまま、他は第一位）
    for col in df_short.select_dtypes(include='number').columns:
        if col != '係数':  # 係数は既に小数点第二位で計算済み
            df_short[col] = df_short[col].apply(my_round)
    
    # --- Sheet 5 ---
    s5 = pd.DataFrame(index=sorted_bases)
    s5.index.name = '拠点名'
    if not df_ext_valid.empty and 'time' in df_ext_valid.columns:
        df_ext_valid_clean = df_ext_valid.dropna(subset=['time'])
        df_ext_biz = df_ext_valid_clean[df_ext_valid_clean['time'].apply(lambda t: BIZ_START <= t <= BIZ_END)]
        s5['営業時間内_外線のみ'] = s5.index.map(df_ext_biz['final_base'].value_counts()).fillna(0).astype(int)
    else: s5['営業時間内_外線のみ'] = 0

    s5['人員'] = s5.index.map(s3.set_index('拠点名')['人員']).fillna(0).astype(int)
    s5['1人当たり／月'] = s5.apply(lambda r: my_round(r['営業時間内_外線のみ'] / r['人員']) if r['人員'] > 0 else 0.0, axis=1)
    total_s5 = s5['営業時間内_外線のみ'].sum()
    s5['全体からの比率'] = s5['営業時間内_外線のみ'].apply(lambda x: x / total_s5 if total_s5 > 0 else 0.0)
    s5 = s5.reset_index()

    # --- Sheet 6 ---
    s6 = pd.DataFrame(index=sorted_bases)
    s6.index.name = '拠点'

    def get_biz_target_counts(df, valid_only=False):
        if df.empty or 'time' not in df.columns or 'target_base' not in df.columns: return {}
        temp_df = df.dropna(subset=['time']).copy()
        filtered = temp_df[temp_df['time'].apply(lambda t: BIZ_START <= t <= BIZ_END)]
        if valid_only:
            filtered = filtered[filtered['final_name'].apply(is_valid_answer)]
        return filtered['target_base'].value_counts()

    s6_data = {
        ('内線', '入電'): s6.index.map(get_biz_target_counts(df_int, valid_only=False)).fillna(0).astype(int),
        ('内線', '着電'): s6.index.map(get_biz_target_counts(df_int, valid_only=True)).fillna(0).astype(int),
        ('外線', '入電'): s6.index.map(get_biz_target_counts(df_ext, valid_only=False)).fillna(0).astype(int),
        ('外線', '着電'): s6.index.map(get_biz_target_counts(df_ext, valid_only=True)).fillna(0).astype(int),
    }
    s6 = pd.DataFrame(s6_data, index=s6.index)
    
    s6[('内線', '応答率')] = s6.apply(lambda r: r[('内線', '着電')]/r[('内線', '入電')] if r[('内線', '入電')]>0 else 0.0, axis=1)
    s6[('外線', '応答率')] = s6.apply(lambda r: r[('外線', '着電')]/r[('外線', '入電')] if r[('外線', '入電')]>0 else 0.0, axis=1)
    s6 = s6[[('内線', '入電'), ('内線', '着電'), ('内線', '応答率'), ('外線', '入電'), ('外線', '着電'), ('外線', '応答率')]]

    s6_total = s6.sum(numeric_only=True)
    s6.loc['合計'] = s6_total
    s6.loc['合計', ('内線', '応答率')] = s6.loc['合計', ('内線', '着電')] / s6.loc['合計', ('内線', '入電')] if s6.loc['合計', ('内線', '入電')] > 0 else 0.0
    s6.loc['合計', ('外線', '応答率')] = s6.loc['合計', ('外線', '着電')] / s6.loc['合計', ('外線', '入電')] if s6.loc['合計', ('外線', '入電')] > 0 else 0.0

    if not df_ext.empty and 'time' in df_ext.columns:
        df_ext_clean = df_ext.dropna(subset=['time']).copy()
        df_ext_biz_all = df_ext_clean[df_ext_clean['time'].apply(lambda t: BIZ_START <= t <= BIZ_END)]
        N1_val = len(df_ext_biz_all)
        df_ext_biz_missed = df_ext_biz_all[~df_ext_biz_all['final_name'].apply(is_valid_answer)]
        O1_val = len(df_ext_biz_missed)
        P1_val = N1_val - O1_val
        Q1_val = P1_val / N1_val if N1_val > 0 else 0.0
        M1_val = s6.loc['合計', ('外線', '入電')]
    else:
        M1_val = 0; N1_val = 0; O1_val = 0; P1_val = 0; Q1_val = 0.0

    s6_header = pd.DataFrame({'表計(入電)': [M1_val], '全入電(N1)': [N1_val], '全不在(O1)': [O1_val], '全着電(P1)': [P1_val], '全応答率(Q1)': [Q1_val]})

    s6_group_rows = [['追加集計(時間内)', '入った数', 'とった数', '％']]
    grp6_total_in = 0
    grp6_total_ans = 0
    for grp in SUMMARY_GROUPS:
        members = [m for m in grp['members'] if m in s6.index and m != '合計']
        if members:
            sum_in = s6.loc[members, ('外線', '入電')].sum()
            sum_ans = s6.loc[members, ('外線', '着電')].sum()
            rate = sum_ans / sum_in if sum_in > 0 else 0.0
            s6_group_rows.append([grp['name'], sum_in, sum_ans, rate])
            grp6_total_in += sum_in
            grp6_total_ans += sum_ans
        else:
            s6_group_rows.append([grp['name'], 0, 0, 0.0])
    grp6_total_rate = grp6_total_ans / grp6_total_in if grp6_total_in > 0 else 0.0
    s6_group_rows.append(['合計', grp6_total_in, grp6_total_ans, grp6_total_rate])
    s6_group_summary = pd.DataFrame(s6_group_rows)

    # 4. 出力
    print("\n集計中...")
    output_filename = f"集計結果_{datetime.date.today()}.xlsx"
    
    try:
        with pd.ExcelWriter(output_filename) as writer:
            # ★レイアウト崩れ対策：index名を一時的に消す
            s1.index.name = None
            s6.index.name = None
            
            s1.to_excel(writer, sheet_name='1.着信件数', startrow=0, startcol=0) 
            s1_summary.to_excel(writer, sheet_name='1.着信件数', index=False, header=False, startrow=21, startcol=9)  # ★J22セルから開始

            s2.to_excel(writer, sheet_name='2.従業員別', index=False)
            s3.to_excel(writer, sheet_name='3.関数_拠点別', index=False)
            # ★受発注グループ集計を右側に出力
            if not s3_juhatchu.empty:
                s3_juhatchu.to_excel(writer, sheet_name='3.関数_拠点別', index=False, startrow=0, startcol=8)
            if not df_short.empty: df_short.to_excel(writer, sheet_name='4.時短勤務', index=False)
            else: pd.DataFrame({'info': ['データなし']}).to_excel(writer, sheet_name='4.時短勤務', index=False)
            
            s5.to_excel(writer, sheet_name='5.営業時間内集計', index=False)
            
            s6.to_excel(writer, sheet_name='6.時間内集計', startrow=3, startcol=0)
            s6_header.to_excel(writer, sheet_name='6.時間内集計', index=False, startrow=0, startcol=12)
            s6_group_summary.to_excel(writer, sheet_name='6.時間内集計', index=False, header=False, startrow=4, startcol=12)

        print(f"\n★★ 完了しました！ ★★")
        print(f"作成されたファイル: {output_filename}")
        
        decorate_excel(output_filename)
        
    except Exception as e:
        print(f"\n【エラー】Excelファイルの保存に失敗しました: {e}")

    input("\nEnterキーを押して終了してください...")

if __name__ == "__main__":
    main()