import pandas as pd
import glob
import re
import os
import sys
import datetime
import warnings

# 警告を無視
warnings.simplefilter('ignore')

# 実行場所をスクリプトのあるフォルダに設定
try:
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
except:
    pass

# ==========================================
# 1. 設定エリア
# ==========================================
BASE_ORDER = ['東京', '横浜', '埼玉', '滋賀', '大阪', '千葉', '福岡', '岡山', '名古屋', '仙台', '流山']
EXCLUDE_KEYWORDS = ['不在', '未応答', '応答なし', '放棄', '留守電']
BIZ_START, BIZ_END = datetime.time(8, 45, 0), datetime.time(17, 45, 0)
TRANS_FROM = {'東京': '流山', '横浜': '多摩', '埼玉': '千葉・北関東', '滋賀': '金沢', '福岡': '広島・熊本', '岡山': '大阪・静岡', '仙台': '札幌'}
TRANS_TO = {'千葉': '埼玉', '大阪': '岡山'}

SUMMARY_GROUPS = [
    {'name': '東京+流山', 'members': ['東京', '流山']},
    {'name': '埼玉+千葉+北関東', 'members': ['埼玉', '千葉', '北関東']},
    {'name': '横浜+多摩', 'members': ['横浜', '多摩']},
    {'name': '仙台+札幌', 'members': ['仙台', '札幌']},
    {'name': '滋賀+金沢', 'members': ['滋賀', '金沢']},
    {'name': '名古屋', 'members': ['名古屋']},
    {'name': '福岡+熊本+広島', 'members': ['福岡', '熊本', '広島']},
    {'name': '岡山+大阪+静岡', 'members': ['岡山', '大阪', '静岡']}
]

JUHATCHU_MEMBERS_RAW = [
    "【流山】鴻池千紗都", "【横浜】奥秋素子", "【横浜】山﨑啓右", "【横浜】青木観奈", "【横浜】川崎瞳", "【横浜】中山満里奈", "【横浜】萩原直美", "【横浜】舞智江",
    "【岡山】伊藤優", "【岡山】河原由布子", "【岡山】岸本麻衣子", "【岡山】岩佐寛子", "【岡山】武政彩果", "【岡山】矢野夏絵", "【埼玉】竹内梓", "【埼玉】吉田夏子",
    "【埼玉】金子友希", "【埼玉】細田昌子", "【埼玉】石川恵理", "【埼玉】池田千恵子", "【埼玉】豊田里花", "【埼玉】矢島静音", "【滋賀】松井治美", "【滋賀】梅原薫",
    "【滋賀】筆坂雪子", "【滋賀】北出智子", "【仙台】阿部かおり", "【仙台】真壁彰子", "【仙台】仁藤佳美", "【仙台】齋藤詩織", "【千葉】小島佳菜", "【大阪】松下愛",
    "【東京】高澤早紀", "【東京】佐々木美咲", "【東京】坂田智世", "【東京】西垣彩", "【東京】北林友希", "【東京】林まど佳", "【福岡】奥薗美和", "【福岡】山崎芹奈",
    "【福岡】重松宝成", "【福岡】松添愛", "【福岡】松尾明日香", "【名古屋】稲垣みちる", "【名古屋】玉腰千恵", "【名古屋】水島奈美", "【名古屋】大渕温子"
]

# ==========================================
# 2. 関数定義
# ==========================================
def extract_base_name(text):
    if not isinstance(text, str): return None, None
    match = re.search(r'【(.*?)】(.*)', text)
    if match: return match.group(1), match.group(2).strip()
    return None, text

def is_valid_answer(name):
    if not isinstance(name, str): return False
    return not any(kw in name for kw in EXCLUDE_KEYWORDS)

def my_round(x):
    try:
        if pd.isna(x): return ""
        val = float(x)
        return int(val * 10 + 0.5) / 10.0
    except: return x

def smart_read_excel(file, sheet_name):
    """ヘッダー位置を自動検出して読み込む"""
    try:
        preview = pd.read_excel(file, sheet_name=sheet_name, header=None, nrows=20)
        target_row = 0
        for i, row in preview.iterrows():
            row_str = row.astype(str).values
            if any('着信者' in s for s in row_str) or any('時刻' in s for s in row_str):
                target_row = i
                break
        return pd.read_excel(file, sheet_name=sheet_name, header=target_row)
    except:
        return pd.DataFrame()

def find_and_load(target_keywords, exclude_keywords=[]):
    """条件に合うファイルを探して読み込む"""
    candidates = []
    # Excel files
    for file in glob.glob("*.xlsx"):
        if "集計結果" in file: continue
        try:
            xls = pd.ExcelFile(file)
            for sheet in xls.sheet_names:
                if any(k in sheet for k in target_keywords):
                    if not any(ek in sheet for ek in exclude_keywords):
                        candidates.append({'file': file, 'sheet': sheet, 'type': 'xlsx'})
        except: pass
    # CSV files
    for file in glob.glob("*.csv"):
        if "集計結果" in file: continue
        if any(k in file for k in target_keywords):
            if not any(ek in file for ek in exclude_keywords):
                candidates.append({'file': file, 'sheet': None, 'type': 'csv'})

    if not candidates: return pd.DataFrame()
    candidates.sort(key=lambda x: (0 if "統合版" in x['file'] else 1))
    t = candidates[0]
    print(f"  -> 読み込み: {os.path.basename(t['file'])} ({t['sheet'] if t['sheet'] else 'CSV'})")
    
    if t['type'] == 'xlsx':
        return smart_read_excel(t['file'], t['sheet'])
    else:
        try: return pd.read_csv(t['file'], encoding='utf-8')
        except: return pd.read_csv(t['file'], encoding='cp932')

# ★デザイン装飾関数（プロフェッショナル版・バグ修正済み）
def decorate_excel(filename):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError: return

    wb = openpyxl.load_workbook(filename)
    FONT_NAME = "BIZ UDゴシック"
    NAVY, PALE_BLUE, LIGHT_YELLOW, WHITE, GRAY = "1F497D", "DCE6F1", "FFF2CC", "FFFFFF", "BFBFBF"
    
    header_fill = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
    header_font = Font(bold=True, color=WHITE, name=FONT_NAME)
    sub_fill = PatternFill(start_color=PALE_BLUE, end_color=PALE_BLUE, fill_type="solid")
    sub_font = Font(bold=True, name=FONT_NAME)
    total_fill = PatternFill(start_color=LIGHT_YELLOW, end_color=LIGHT_YELLOW, fill_type="solid")
    border = Border(left=Side(style='thin', color=GRAY), right=Side(style='thin', color=GRAY), top=Side(style='thin', color=GRAY), bottom=Side(style='thin', color=GRAY))
    base_font = Font(name=FONT_NAME)
    note_font = Font(name=FONT_NAME, color="FF0000", size=10, bold=True)

    def apply_style(ws, r1, c1, r2, c2, mode="data"):
        """範囲にスタイルを適用 (行・列の順序: StartRow, StartCol, EndRow, EndCol)"""
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = border; cell.font = base_font
                if isinstance(cell.value, (int, float)):
                    # 数値書式
                    val_str = str(ws.cell(row=r1, column=c).value) # ヘッダー参照
                    if any(x in val_str for x in ["率","％","比"]): cell.number_format = '0.0%'
                    else: cell.number_format = '#,##0'
                
                if mode == "header":
                    cell.fill = header_fill; cell.font = header_font; cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                elif mode == "sub":
                    cell.fill = sub_fill; cell.font = sub_font; cell.alignment = Alignment(horizontal='center', vertical='center')
                elif mode == "total":
                    cell.fill = total_fill; cell.font = Font(bold=True, name=FONT_NAME)
                elif mode == "data":
                    cell.fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

    def clean_area(ws, r1, c1, r2, c2):
        """範囲をクリア（余白用）"""
        for r in range(r1, r2 + 1):
            for c in range(c1, c2 + 1):
                cell = ws.cell(row=r, column=c)
                cell.fill = PatternFill(fill_type=None); cell.border = None; cell.value = ""

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]; ws.sheet_view.showGridLines = False
        # 全セルフォント適用
        for row in ws.iter_rows():
            for cell in row: cell.font = base_font
        
        # 列幅調整
        for col in ws.columns:
            max_l = 0
            for cell in col:
                try: max_l = max(max_l, len(str(cell.value)))
                except: pass
            ws.column_dimensions[get_column_letter(col[0].column)].width = min((max_l + 2) * 1.1, 50)

        mr = ws.max_row
        
        if sheet_name == '1.着信件数':
            # Main Table (A1:G) Header: Row 1-2
            apply_style(ws, 1, 1, 2, 7, "header")
            ws.merge_cells('A1:A2'); ws.merge_cells('B1:C1'); ws.merge_cells('D1:E1'); ws.merge_cells('F1:F2'); ws.merge_cells('G1:G2')
            ws['A1'] = "拠点"
            apply_style(ws, 3, 1, mr-1, 7, "data"); apply_style(ws, mr, 1, mr, 7, "total")
            
            # H列クリア
            clean_area(ws, 1, 8, mr+5, 9); ws.column_dimensions['H'].width = 2
            
            # J列 (Sub Tables)
            ws['J12'].font = Font(bold=True, size=11, name=FONT_NAME)
            apply_style(ws, 13, 11, 13, 13, "sub"); ws['J13'].border = border # K-M Header
            clean_area(ws, 13, 10, 13, 10); ws['J13'].border = border # J13 Clean but border
            apply_style(ws, 14, 10, 15, 13, "data")
            
            # Group Summary
            for row in ws.iter_rows(min_col=10, max_col=10):
                if row[0].value == "追加集計(24H)":
                    sr = row[0].row; er = sr + len(SUMMARY_GROUPS) + 2
                    apply_style(ws, sr, 10, sr+1, 13, "sub")
                    apply_style(ws, sr+2, 10, er-1, 13, "data")
                    apply_style(ws, er, 10, er, 13, "total")

        elif sheet_name == '5.営業時間内集計':
            ws.insert_rows(1); ws['A1'] = "※集計基準：『誰が取ったか（Final Base）』でカウント"; ws['A1'].font = note_font
            apply_style(ws, 2, 1, 2, ws.max_column, "header")
            if ws.max_row >= 3: apply_style(ws, 3, 1, ws.max_row, ws.max_column, "data")

        elif sheet_name == '6.時間内集計':
            ws.insert_rows(1); ws['A1'] = "※集計基準：『どこ宛てか（Target Base）』でカウント"; ws['A1'].font = note_font
            mr = ws.max_row
            # Main Table (Header: 4-5)
            apply_style(ws, 4, 1, 5, 7, "header")
            ws.merge_cells('A4:A5'); ws.merge_cells('B4:D4'); ws.merge_cells('E4:G4'); ws['A4'] = "拠点"
            apply_style(ws, 6, 1, mr-1, 7, "data"); apply_style(ws, mr, 1, mr, 7, "total")
            
            # Clear Middle
            clean_area(ws, 1, 8, mr+5, 12)
            
            # Top Summary (M2:Q3)
            apply_style(ws, 2, 13, 2, 17, "sub"); apply_style(ws, 3, 13, 3, 17, "data")
            for c in range(13, 18): ws.cell(row=3, column=c).font = Font(bold=True, name=FONT_NAME)
            
            # Group Summary (M6)
            sr = 6; er = sr + len(SUMMARY_GROUPS) + 2 + 1
            apply_style(ws, sr, 13, sr+1, 16, "sub")
            apply_style(ws, sr+2, 13, er-1, 16, "data"); apply_style(ws, er, 13, er, 16, "total")

        else: # Other sheets
            if ws.max_row > 0: apply_style(ws, 1, 1, 1, ws.max_column, "header")
            if ws.max_row > 1: apply_style(ws, 2, 1, ws.max_row, ws.max_column, "data")

    wb.save(filename)

# ==========================================
# 3. メイン処理
# ==========================================
def main():
    print(f"作業フォルダ: {os.getcwd()}")
    print("データを探しています...")

    df_int = find_and_load(['内線通話', '内線'], exclude_keywords=['発信'])
    df_ext = find_and_load(['外線着信', '外線'], exclude_keywords=['発信'])
    df_short = find_and_load(['時短勤務', '時短'])

    print(f"\nデータ件数: 内線={len(df_int)}件, 外線={len(df_ext)}件, 時短={len(df_short)}件")

    if df_int.empty and df_ext.empty:
        print("\n【エラー】データが見つかりません。")
        return

    # 年月取得と基本処理
    data_year_month = None
    for df in [df_int, df_ext]:
        if df.empty: continue
        df.columns = [str(c).strip() for c in df.columns]
        if '着信者' in df.columns: df['target_base'], _ = zip(*df['着信者'].apply(extract_base_name))
        if '最終着信者' in df.columns: df['final_base'], df['final_name'] = zip(*df['最終着信者'].apply(extract_base_name))
        else: df['final_base'] = None; df['final_name'] = None
        if '時刻' in df.columns:
            df['dt'] = pd.to_datetime(df['時刻'].astype(str), errors='coerce')
            df['day'] = df['dt'].dt.day
            df['time'] = df['dt'].dt.time
            df['date'] = df['dt'].dt.date
            if data_year_month is None and not df['dt'].dropna().empty:
                f_dt = df['dt'].dropna().iloc[0]
                data_year_month = (f_dt.year, f_dt.month)

    all_dates = set()
    if not df_int.empty: all_dates.update(df_int['date'].dropna().unique())
    if not df_ext.empty: all_dates.update(df_ext['date'].dropna().unique())
    total_operating_days = len(all_dates) if all_dates else 1
    
    # フィルタリング済みDF作成
    df_int_v = df_int[df_int['final_name'].apply(is_valid_answer)] if not df_int.empty else pd.DataFrame()
    df_ext_v = df_ext[df_ext['final_name'].apply(is_valid_answer)] if not df_ext.empty else pd.DataFrame()

    # --- Sheet 1 (着信件数) ---
    all_bases = set(BASE_ORDER)
    if not df_int.empty: all_bases |= set(df_int['target_base'].dropna())
    if not df_ext.empty: all_bases |= set(df_ext['target_base'].dropna())
    sorted_bases = sorted(list(all_bases), key=lambda x: BASE_ORDER.index(x) if x in BASE_ORDER else 999)

    s1 = pd.DataFrame(index=sorted_bases); s1.index.name = '拠点'
    def cnt(df, col): return df[col].value_counts() if not df.empty else {}

    s1_data = {
        ('内線','入電'): s1.index.map(cnt(df_int,'target_base')).fillna(0), ('内線','着電'): s1.index.map(cnt(df_int_v,'final_base')).fillna(0),
        ('外線','入電'): s1.index.map(cnt(df_ext,'target_base')).fillna(0), ('外線','着電'): s1.index.map(cnt(df_ext_v,'final_base')).fillna(0),
        ('他拠点へ転送',' '): s1.index.map(TRANS_TO).fillna(''), ('他拠点から転送',' '): s1.index.map(TRANS_FROM).fillna('')
    }
    s1 = pd.DataFrame(s1_data, index=s1.index); s1.loc['合計'] = s1.sum(numeric_only=True)
    s1.loc['合計', ('他拠点へ転送',' ')] = ''; s1.loc['合計', ('他拠点から転送',' ')] = ''

    # S1 Group Summary
    s1_g = [['追加集計(24H)', '電話が入った数', 'とった数', '％']]; gi, ga = 0, 0
    for grp in SUMMARY_GROUPS:
        mems = [m for m in grp['members'] if m in s1.index and m != '合計']
        si, sa = s1.loc[mems, ('外線','入電')].sum(), s1.loc[mems, ('外線','着電')].sum()
        s1_g.append([grp['name'], si, sa, sa/si if si>0 else 0]); gi += si; ga += sa
    s1_g.append(['合計', gi, ga, ga/gi if gi>0 else 0])

    # --- Sheet 2 (従業員別) ---
    s2 = pd.DataFrame()
    if not df_int_v.empty or not df_ext_v.empty:
        s2_i = df_int_v.groupby(['final_name','final_base']).agg(内線=('通話時間','size'), 内線秒=('通話時間','sum')) if not df_int_v.empty else pd.DataFrame(columns=['内線','内線秒'])
        s2_e = df_ext_v.groupby(['final_name','final_base']).agg(外線=('通話時間','size'), 外線秒=('通話時間','sum')) if not df_ext_v.empty else pd.DataFrame(columns=['外線','外線秒'])
        s2 = s2_i.join(s2_e, how='outer').fillna(0)
        
        j_set = set(tuple(extract_base_name(x)) for x in JUHATCHU_MEMBERS_RAW)
        s2['受発注'] = s2.index.map(lambda x: "受発注" if (x[1], x[0]) in j_set else "")
        s2['内線件数'] = s2['内線']
        s2['通話時間／秒'] = (s2['内線秒']/s2['内線'].replace(0,1)).apply(my_round)
        s2['外線件数'] = s2['外線']
        s2['通話時間／秒.1'] = (s2['外線秒']/s2['外線'].replace(0,1)).apply(my_round)
        
        df_all = pd.concat([df_int_v, df_ext_v])
        
        # ★日付データのクリーニング
        if not df_all.empty:
            df_all = df_all.dropna(subset=['day']) # 空欄の日付を削除
            df_all['day'] = df_all['day'].astype(int) # 整数型に変換
            
            if data_year_month:
                y, m = data_year_month
                daily = df_all.pivot_table(index=['final_name','final_base'], columns='day', aggfunc='size', fill_value=0)
                # 日付列の生成
                date_cols = {}
                for d in daily.columns:
                    try: date_cols[d] = datetime.date(y, m, int(d)).strftime('%Y-%m-%d')
                    except: pass
                daily = daily.rename(columns=date_cols)
                s2 = s2.join(daily).reset_index().rename(columns={'final_name':'名前','final_base':'拠点'})
                s2['内外線計'] = s2['内線'] + s2['外線']; s2['稼働日'] = (daily > 0).sum(axis=1).values
                s2['1日平均'] = (s2['内外線計']/s2['稼働日'].replace(0,1)).apply(my_round)
                s2 = s2.sort_values(['拠点','名前'])
                final_cols = ['名前','拠点','受発注','内線件数','通話時間／秒','外線件数','通話時間／秒.1'] + sorted(list(date_cols.values())) + ['稼働日','内外線計','1日平均']
                s2 = s2[[c for c in final_cols if c in s2.columns]]

    # --- Sheet 3 & 5 ---
    s3 = s1[[('外線','着電')]].reset_index(); s3.columns = ['拠点名','外線計']; s3 = s3[s3['拠点名'] != '合計']
    nm = f"{data_year_month[0]}年{data_year_month[1]}月から外線のみ" if data_year_month else "期間計"
    s3 = s3.rename(columns={'外線計': nm})
    s3['外線のみ'] = s3[nm].apply(lambda x: my_round(x/total_operating_days))
    s3['人員'] = s3['拠点名'].map(s2.groupby('拠点')['名前'].nunique()).fillna(0)
    s3['1人当たり／月'] = s3.apply(lambda r: my_round(r[nm]/r['人員']) if r['人員']>0 else 0, axis=1)
    s3['全体からの比率'] = s3[nm] / s3[nm].sum()

    s5 = pd.DataFrame(index=sorted_bases); t_biz = df_ext_v[df_ext_v['time'].apply(lambda t: BIZ_START <= t <= BIZ_END)] if not df_ext_v.empty else pd.DataFrame()
    s5['営業時間内_外線のみ'] = s5.index.map(t_biz['final_base'].value_counts()).fillna(0)
    s5['人員'] = s5.index.map(s3.set_index('拠点名')['人員']).fillna(0)
    s5['1人当たり／月'] = s5.apply(lambda r: my_round(r['営業時間内_外線のみ']/r['人員']) if r['人員']>0 else 0, axis=1)
    s5['全体からの比率'] = s5['営業時間内_外線のみ'] / s5['営業時間内_外線のみ'].sum()
    s5 = s5.reset_index().rename(columns={'index':'拠点名'})

    # --- Sheet 6 ---
    def get_biz(df, v=False):
        if df.empty: return {}
        t = df.dropna(subset=['time']); t = t[t['time'].apply(lambda x: BIZ_START <= x <= BIZ_END)]
        if v: t = t[t['final_name'].apply(is_valid_answer)]
        return t['target_base'].value_counts()
    s6 = pd.DataFrame(index=sorted_bases)
    s6[('内線','入電')], s6[('内線','着電')] = s6.index.map(get_biz(df_int)).fillna(0), s6.index.map(get_biz(df_int,True)).fillna(0)
    s6[('外線','入電')], s6[('外線','着電')] = s6.index.map(get_biz(df_ext)).fillna(0), s6.index.map(get_biz(df_ext,True)).fillna(0)
    s6[('内線','応答率')] = s6[('内線','着電')] / s6[('内線','入電')].replace(0,1)
    s6[('外線','応答率')] = s6[('外線','着電')] / s6[('外線','入電')].replace(0,1)
    s6 = s6[[('内線','入電'),('内線','着電'),('内線','応答率'),('外線','入電'),('外線','着電'),('外線','応答率')]]
    s6.loc['合計'] = s6.sum(); s6.loc['合計',('内線','応答率')] = s6.loc['合計',('内線','着電')]/s6.loc['合計',('内線','入電')]
    s6.loc['合計',('外線','応答率')] = s6.loc['合計',('外線','着電')]/s6.loc['合計',('外線','入電')]
    
    t6 = df_ext.dropna(subset=['time']); t6 = t6[t6['time'].apply(lambda x: BIZ_START <= x <= BIZ_END)]
    n1 = len(t6); o1 = len(t6[~t6['final_name'].apply(is_valid_answer)]); m1 = s6.loc['合計',('外線','入電')]
    s6_h = pd.DataFrame({'表計(入電)':[m1],'全入電(N1)':[n1],'全不在(O1)':[o1],'全着電(P1)':[n1-o1],'全応答率(Q1)':[(n1-o1)/n1 if n1>0 else 0]})
    s6_g = [['追加集計(時間内)', '入った数', 'とった数', '％']]; gi, ga = 0, 0
    for grp in SUMMARY_GROUPS:
        mems = [m for m in grp['members'] if m in s6.index and m != '合計']
        si, sa = s6.loc[mems, ('外線','入電')].sum(), s6.loc[mems, ('外線','着電')].sum()
        s6_g.append([grp['name'], si, sa, sa/si if si>0 else 0]); gi += si; ga += sa
    s6_g.append(['合計', gi, ga, ga/gi if gi>0 else 0])

    # 4. 出力
    print("\n集計中...")
    output_filename = f"集計結果_{datetime.date.today()}.xlsx"
    try:
        with pd.ExcelWriter(output_filename) as writer:
            s1.to_excel(writer, '1.着信件数'); pd.DataFrame(s1_g).to_excel(writer, '1.着信件数', index=False, header=False, startrow=11, startcol=9)
            s2.to_excel(writer, '2.従業員別', index=False); s3.to_excel(writer, '3.関数_拠点別', index=False)
            (df_short if not df_short.empty else pd.DataFrame({'info':['データなし']})).to_excel(writer, '4.時短勤務', index=False)
            s5.to_excel(writer, '5.営業時間内集計', index=False); s6.to_excel(writer, '6.時間内集計', startrow=3)
            s6_h.to_excel(writer, '6.時間内集計', index=False, startrow=0, startcol=12); pd.DataFrame(s6_g).to_excel(writer, '6.時間内集計', index=False, header=False, startrow=4, startcol=12)
        
        print(f"\n★★ 完了しました！ ★★")
        print(f"作成されたファイル: {output_filename}")
        decorate_excel(output_filename)
    except Exception as e:
        print(f"\n【エラー】Excelファイルの保存に失敗しました: {e}")

    input("\nEnterキーを押して終了してください...")

if __name__ == "__main__":
    main()